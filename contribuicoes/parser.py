import os
import re
from datetime import datetime
from collections import OrderedDict

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook


def ler_definicoes_registros(arquivo_markdown):
    """
    Lê o arquivo markdown e extrai as definições de todos os registros.
    
    Args:
        arquivo_markdown: Caminho para o arquivo sped_bloco_0.md
        
    Returns:
        Dicionário com código do registro como chave e lista de campos como valor
        Exemplo: {'0000': [{'numero': '01', 'campo': 'REG', 'tipo': 'C', ...}, ...]}
    """
    definicoes = OrderedDict()
    
    if not os.path.exists(arquivo_markdown):
        print(f"Erro: Arquivo '{arquivo_markdown}' não encontrado!")
        return definicoes
    
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_markdown, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding do markdown detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo markdown!")
        return definicoes
    
    try:
        conteudo = arquivo.read()
    finally:
        arquivo.close()
    
    padrao_registro = r'##\s+Registro\s+(\d+):'

    matches = list(re.finditer(padrao_registro, conteudo, re.IGNORECASE))
    
    for i, match in enumerate(matches):
        codigo_registro = match.group(1).zfill(4)
        inicio = match.end()
        
        if i + 1 < len(matches):
            fim = matches[i + 1].start()
        else:
            fim = len(conteudo)
        
        secao = conteudo[inicio:fim]
        
        padroes_tabela = [
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[çc][ãa]o\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descricao\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[çc]ao\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[ãa]o\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
        ]
        
        match_tabela = None
        for padrao in padroes_tabela:
            match_tabela = re.search(padrao, secao, re.IGNORECASE)
            if match_tabela:
                break
        
        if match_tabela:
            inicio_tabela = match_tabela.end()
            linhas = secao[inicio_tabela:].split('\n')
            
            campos = []
            encontrou_separador = False
            
            for idx, linha in enumerate(linhas):
                linha = linha.strip()
                
                if linha.startswith('|') and '----' in linha:
                    encontrou_separador = True
                    continue
                
                if (linha.startswith('###') or linha.startswith('Observações') or 
                    linha.startswith('Observacoes') or linha.startswith('Nível') or 
                    linha.startswith('Nivel') or linha.startswith('# ')):
                    if campos:
                        break
                    continue
                
                if not encontrou_separador:
                    continue
                
                if not linha.startswith('|'):
                    partes_espaco = linha.split(None, 2)
                    if len(partes_espaco) >= 2 and partes_espaco[0].isdigit():
                        numero_campo = partes_espaco[0]
                        resto_linha = partes_espaco[1] if len(partes_espaco) > 1 else ''
                        resto_descricao = partes_espaco[2] if len(partes_espaco) > 2 else ''
                        
                        palavras = resto_linha.split()
                        nome_campo = palavras[0] if palavras else ''
                        
                        descricao_completa = resto_descricao if resto_descricao else ''
                        if not descricao_completa and len(palavras) > 1:
                            descricao_completa = ' '.join(palavras[1:])
                        
                        padrao_tipo_tam = r'([CN])\s+(\d+\*?)\s*-\s*([SN])'
                        match_tipo = re.search(padrao_tipo_tam, linha)
                        
                        if match_tipo:
                            tipo = match_tipo.group(1)
                            tamanho = match_tipo.group(2)
                            obrigatorio = match_tipo.group(3)
                            
                            descricao = re.sub(padrao_tipo_tam, '', linha).strip()
                            descricao = descricao.replace(numero_campo, '', 1).strip()
                            descricao = descricao.replace(nome_campo, '', 1).strip()
                            descricao = descricao.rstrip(':').strip()
                            
                            campo = {
                                'numero': numero_campo,
                                'campo': nome_campo,
                                'descricao': descricao,
                                'tipo': tipo,
                                'tamanho': tamanho,
                                'decimais': '',
                                'obrigatorio': obrigatorio
                            }
                            campos.append(campo)
                            continue
                    
                    if campos:
                        ultimo_campo = campos[-1]
                        
                        texto_continuacao = linha.strip()
                        if texto_continuacao and not texto_continuacao.startswith('#') and not (len(texto_continuacao) > 0 and texto_continuacao[0].isdigit()):
                            ultimo_campo['descricao'] += ' ' + texto_continuacao
                    continue
                
                if linha.startswith('|') and linha.count('|') >= 3:
                    partes = [p.strip() for p in linha.split('|')]

                    if partes and not partes[0]:
                        partes = partes[1:]

                    if partes and not partes[-1]:
                        partes = partes[:-1]
                    
                    if len(partes) >= 7:
                        try:
                            if partes[0].isdigit():
                                nome_campo = partes[1]
                                
                                idx_tipo = None
                                for i in range(2, len(partes)):
                                    if partes[i] and partes[i][0] in ['C', 'N']:
                                        idx_tipo = i
                                        break
                                
                                if idx_tipo is None:
                                    continue
                                
                                descricao = ' '.join(partes[2:idx_tipo]).strip()
                                
                                tipo_completo = partes[idx_tipo]
                                tipo = tipo_completo.split()[0] if tipo_completo else ''
                                
                                tamanho_completo = partes[idx_tipo + 1] if idx_tipo + 1 < len(partes) else ''
                                tamanho = tamanho_completo.split()[0] if tamanho_completo else ''
                                
                                decimais_completo = partes[idx_tipo + 2] if idx_tipo + 2 < len(partes) else ''
                                decimais = decimais_completo.split()[0] if decimais_completo else ''
                                
                                obrigatorio_completo = partes[-1] if partes else ''
                                obrigatorio = obrigatorio_completo.split()[0] if obrigatorio_completo else ''
                                
                                campo = {
                                    'numero': partes[0],
                                    'campo': nome_campo,
                                    'descricao': descricao,
                                    'tipo': tipo,
                                    'tamanho': tamanho,
                                    'decimais': decimais,
                                    'obrigatorio': obrigatorio
                                }
                                campos.append(campo)
                        except (IndexError, ValueError) as e:
                            continue
            
            if campos:
                definicoes[codigo_registro] = campos
                print(f"Registro {codigo_registro} encontrado com {len(campos)} campos")
    
    print(f"\nTotal de registros encontrados: {len(definicoes)}")
    return definicoes


def parse_registro_generico(linha, codigo_registro, campos_definicao):
    """
    Parseia uma linha do SPED de forma genérica baseado nas definições.
    
    Args:
        linha: String com a linha do arquivo SPED
        codigo_registro: Código do registro (ex: '0000')
        campos_definicao: Lista de dicionários com as definições dos campos
        
    Returns:
        Dicionário com os campos parseados ou None se inválido
    """
    linha = linha.strip()
    if not linha or codigo_registro not in linha:
        return None
    
    partes = linha.split('|')
    
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == codigo_registro:
                idx = i
                break
        
        if idx is None:
            return None
        
        campos = {}
        for campo_def in campos_definicao:
            numero_campo = int(campo_def['numero'])
            nome_campo = campo_def['campo']
            
            posicao = idx + (numero_campo - 1)
            
            if posicao < len(partes):
                valor = partes[posicao].strip()
                campos[nome_campo] = valor
            else:
                campos[nome_campo] = ''
        
        return campos
        
    except (IndexError, ValueError, KeyError) as e:
        print(f"Erro ao parsear linha do registro {codigo_registro}: {e}")
        return None


def formatar_cnpj(cnpj_str):
    """
    Formata CNPJ para o formato XX.XXX.XXX/XXXX-XX.
    
    Args:
        cnpj_str: String com CNPJ sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cnpj_str or len(cnpj_str) != 14:
        return cnpj_str
    
    try:
        return f"{cnpj_str[0:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:14]}"
    except:
        return cnpj_str


def formatar_cpf(cpf_str):
    """
    Formata CPF para o formato XXX.XXX.XXX-XX.
    
    Args:
        cpf_str: String com CPF sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cpf_str or len(cpf_str) != 11:
        return cpf_str
    
    try:
        return f"{cpf_str[0:3]}.{cpf_str[3:6]}.{cpf_str[6:9]}-{cpf_str[9:11]}"
    except:
        return cpf_str


def formatar_cep(cep_str):
    """
    Formata CEP para o formato XXXXX-XXX.
    
    Args:
        cep_str: String com CEP sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cep_str or len(cep_str) != 8:
        return cep_str
    
    try:
        return f"{cep_str[0:5]}-{cep_str[5:8]}"
    except:
        return cep_str


def formatar_telefone(tel_str):
    """
    Formata telefone para o formato (XX) XXXXX-XXXX ou (XX) XXXX-XXXX.
    
    Args:
        tel_str: String com telefone sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not tel_str:
        return tel_str
    
    try:
        tel_limpo = ''.join(filter(str.isdigit, tel_str))
        if len(tel_limpo) == 10:
            return f"({tel_limpo[0:2]}) {tel_limpo[2:6]}-{tel_limpo[6:10]}"
        elif len(tel_limpo) == 11:
            return f"({tel_limpo[0:2]}) {tel_limpo[2:7]}-{tel_limpo[7:11]}"
        else:
            return tel_str
    except:
        return tel_str


def formatar_data(data_str):
    """
    Formata data do formato DDMMYYYY para DD/MM/YYYY.
    
    Args:
        data_str: String com data no formato DDMMYYYY
        
    Returns:
        String formatada como DD/MM/YYYY ou a string original se inválida
    """
    if not data_str or len(data_str) != 8:
        return data_str
    
    try:
        return f"{data_str[0:2]}/{data_str[2:4]}/{data_str[4:8]}"
    except:
        return data_str


def formatar_valor_monetario(valor_str):
    """
    Formata valor monetário substituindo vírgula por ponto para compatibilidade com Excel.
    
    Args:
        valor_str: String com valor monetário (formato brasileiro com vírgula)
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not valor_str:
        return valor_str
    
    try:
        return valor_str.replace(',', '.')
    except:
        return valor_str


def obter_descricao_campo(campo, valor):
    """
    Retorna descrição formatada para campos específicos baseado no valor.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        
    Returns:
        String com descrição formatada ou valor original
    """
    if not valor:
        return valor
    
    descricoes = {
        'IND_MOV': {
            '0': '0 - Bloco com dados informados',
            '1': '1 - Bloco sem dados informados'
        },
        'TIPO_ESCRIT': {
            '0': '0 - Original',
            '1': '1 - Retificadora'
        },
        'IND_SIT_ESP': {
            '0': '0 - Abertura',
            '1': '1 - Cisão',
            '2': '2 - Fusão',
            '3': '3 - Incorporação',
            '4': '4 - Encerramento'
        },
        'IND_NAT_PJ': {
            '00': '00 - Pessoa jurídica em geral',
            '01': '01 - Sociedade cooperativa',
            '02': '02 - Entidade sujeita ao PIS/Pasep exclusivamente com base na Folha de Salários',
            '03': '03 - Pessoa jurídica em geral participante de SCP como sócia ostensiva',
            '04': '04 - Sociedade cooperativa participante de SCP como sócia ostensiva',
            '05': '05 - Sociedade em Conta de Participação - SCP'
        },
        'IND_ATIV': {
            '0': '0 - Industrial ou equiparado a industrial',
            '1': '1 - Prestador de serviços',
            '2': '2 - Atividade de comércio',
            '3': '3 - Pessoas jurídicas referidas nos §§ 6º, 8º e 9º do art. 3º da Lei nº 9.718, de 1998',
            '4': '4 - Atividade imobiliária',
            '9': '9 - Outros'
        },
        'COD_INC_TRIB': {
            '1': '1 - Escrituração de operações com incidência exclusivamente no regime não-cumulativo',
            '2': '2 - Escrituração de operações com incidência exclusivamente no regime cumulativo',
            '3': '3 - Escrituração de operações com incidência nos regimes não-cumulativo e cumulativo'
        },
        'IND_APRO_CRED': {
            '1': '1 - Método de Apropriação Direta',
            '2': '2 - Método de Rateio Proporcional (Receita Bruta)'
        },
        'COD_TIPO_CONT': {
            '1': '1 - Apuração da Contribuição Exclusivamente a Alíquota Básica',
            '2': '2 - Apuração da Contribuição a Alíquotas Específicas (Diferenciadas e/ou por Unidade de Medida de Produto)'
        },
        'IND_REG_CUM': {
            '1': '1 - Regime de Caixa – Escrituração consolidada (Registro F500)',
            '2': '2 - Regime de Competência - Escrituração consolidada (Registro F550)',
            '9': '9 - Regime de Competência - Escrituração detalhada, com base nos registros dos Blocos "A", "C", "D" e "F"'
        },
        'COD_NAT_CC': {
            '01': '01 - Contas de ativo',
            '02': '02 - Contas de passivo',
            '03': '03 - Patrimônio líquido',
            '04': '04 - Contas de resultado',
            '05': '05 - Contas de compensação',
            '09': '09 - Outras'
        },
        'IND_CTA': {
            'S': 'S - Sintética (grupo de contas)',
            'A': 'A - Analítica (conta)'
        }
    }
    
    if campo in descricoes and valor in descricoes[campo]:
        return descricoes[campo][valor]
    
    return valor


def aplicar_formatacao(campo, valor, tipo_campo):
    """
    Aplica formatação específica baseada no tipo e nome do campo.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        tipo_campo: Tipo do campo (C, N, etc)
        
    Returns:
        Valor formatado
    """
    if not valor:
        return valor
    
    if campo == 'CNPJ' and len(valor) == 14:
        return formatar_cnpj(valor)
    
    if campo == 'CPF' and len(valor) == 11:
        return formatar_cpf(valor)
    
    if campo == 'CEP' and len(valor) == 8:
        return formatar_cep(valor)
    
    if campo in ['FONE', 'FAX']:
        return formatar_telefone(valor)
    
    if campo.startswith('DT_') or campo.startswith('DTE_'):
        if len(valor) == 8 and valor.isdigit():
            return formatar_data(valor)
    
    if campo.startswith('VL_') or campo.startswith('REC_') and tipo_campo == 'N':
        return formatar_valor_monetario(valor)
    
    return obter_descricao_campo(campo, valor)


def ler_arquivo_registro(caminho_arquivo, codigo_registro, campos_definicao):
    """
    Lê o arquivo de um registro específico e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo do registro
        codigo_registro: Código do registro (ex: '0000')
        campos_definicao: Lista de dicionários com as definições dos campos
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo '{caminho_arquivo}' não encontrado (pode não existir registros deste tipo)")
        return registros
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo '{caminho_arquivo}'!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_generico(linha, codigo_registro, campos_definicao)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo=None):
    """
    Cria um arquivo Excel (.xlsx) com os registros parseados.
    
    Args:
        registros: Lista de dicionários com os registros
        codigo_registro: Código do registro (ex: '0000')
        campos_definicao: Lista de dicionários com as definições dos campos
        nome_arquivo: Nome do arquivo Excel a ser criado (opcional)
    """
    if not registros:
        print(f"Nenhum registro {codigo_registro} encontrado!")
        return None
    
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"registro_{codigo_registro}_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {codigo_registro}"
    
    cabecalhos = [campo['campo'] for campo in campos_definicao]
    
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    linha_atual = 2
    
    for registro in registros:
        valores = []
        for campo_def in campos_definicao:
            nome_campo = campo_def['campo']
            tipo_campo = campo_def['tipo']
            valor = registro.get(nome_campo, '')
            
            valor_formatado = aplicar_formatacao(nome_campo, valor, tipo_campo)
            valores.append(valor_formatado if valor_formatado else '')
        
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor)
        
        linha_atual += 1
    
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    caminho_markdown = "sped_bloco_0.md"
    
    print("=" * 60)
    print("Parser Genérico do Bloco 0 - SPED EFD Contribuições")
    print("=" * 60)
    print(f"\nLendo definições do arquivo: {caminho_markdown}")
    
    definicoes = ler_definicoes_registros(caminho_markdown)
    
    if not definicoes:
        print("\nErro: Nenhuma definição de registro encontrada no arquivo markdown!")
        exit(1)
    
    print(f"\n{len(definicoes)} registros encontrados no markdown")
    print("\nProcessando arquivos...")
    print("=" * 60)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    for codigo_registro, campos_definicao in definicoes.items():
        caminho_arquivo = f"blocos_separados/bloco_0/{codigo_registro}.txt"
        
        print(f"\nProcessando registro {codigo_registro}...")
        registros = ler_arquivo_registro(caminho_arquivo, codigo_registro, campos_definicao)
        
        if registros:
            nome_arquivo_excel = f"registro_{codigo_registro}_{timestamp}.xlsx"
            criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo_excel)
        else:
            print(f"Nenhum registro {codigo_registro} encontrado nos arquivos de entrada")
    
    print("\n" + "=" * 60)
    print("Processamento concluído!")
    print("=" * 60)

import os
from datetime import datetime

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook


def parse_registro_A001(linha):
    """
    Parseia uma linha do registro A001 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A001 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A001' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A001'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A001':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A001
        # Estrutura: |A001|IND_MOV|
        # Após split por '|': ['', 'A001', 'IND_MOV', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'IND_MOV': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def obter_descricao_ind_mov(valor):
    """
    Retorna a descrição do indicador de movimento.
    
    Args:
        valor: Valor do campo IND_MOV
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Bloco com dados informados',
        '1': '1 - Bloco sem dados'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def ler_arquivo_A001(caminho_arquivo):
    """
    Lê o arquivo A001.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A001.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A001(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A001(registros, nome_arquivo="registro_A001.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A001 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A001 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A001"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'IND_MOV'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            obter_descricao_ind_mov(registro.get('IND_MOV', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def formatar_cnpj(cnpj_str):
    """
    Formata CNPJ para o formato XX.XXX.XXX/XXXX-XX.
    
    Args:
        cnpj_str: String com CNPJ sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cnpj_str or len(cnpj_str) != 14:
        return cnpj_str
    
    try:
        return f"{cnpj_str[0:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:14]}"
    except:
        return cnpj_str


def parse_registro_A010(linha):
    """
    Parseia uma linha do registro A010 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A010 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A010' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A010'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A010':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A010
        # Estrutura: |A010|CNPJ|
        # Após split por '|': ['', 'A010', 'CNPJ', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'CNPJ': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_A010(caminho_arquivo):
    """
    Lê o arquivo A010.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A010.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A010(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A010(registros, nome_arquivo="registro_A010.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A010 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A010 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A010"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'CNPJ'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        cnpj_formatado = formatar_cnpj(registro.get('CNPJ', ''))
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            cnpj_formatado,
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def formatar_data(data_str):
    """
    Formata data do formato DDMMYYYY para DD/MM/YYYY.
    
    Args:
        data_str: String com data no formato DDMMYYYY
        
    Returns:
        String formatada como DD/MM/YYYY ou a string original se inválida
    """
    if not data_str or len(data_str) != 8:
        return data_str
    
    try:
        return f"{data_str[0:2]}/{data_str[2:4]}/{data_str[4:8]}"
    except:
        return data_str


def formatar_valor_monetario(valor_str):
    """
    Formata valor monetário substituindo vírgula por ponto para compatibilidade com Excel.
    
    Args:
        valor_str: String com valor monetário (formato brasileiro com vírgula)
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not valor_str:
        return valor_str
    
    try:
        # Substitui vírgula por ponto para compatibilidade com Excel
        return valor_str.replace(',', '.')
    except:
        return valor_str


def parse_registro_A100(linha):
    """
    Parseia uma linha do registro A100 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A100 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A100' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A100'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A100':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A100
        # Estrutura: |A100|IND_OPER|IND_EMIT|COD_PART|COD_SIT|SER|SUB|NUM_DOC|CHV_NFE|DT_DOC|DT_EXE_SERV|VL_DOC|IND_PGTO|VL_DESC|VL_BC_PIS|VL_PIS|VL_BC_COFINS|VL_COFINS|VL_PIS_RET|VL_COFINS_RET|VL_ISS|
        # Após split por '|': ['', 'A100', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_SIT', 'SER', 'SUB', 'NUM_DOC', 'CHV_NFE', 'DT_DOC', 'DT_EXE_SERV', 'VL_DOC', 'IND_PGTO', 'VL_DESC', 'VL_BC_PIS', 'VL_PIS', 'VL_BC_COFINS', 'VL_COFINS', 'VL_PIS_RET', 'VL_COFINS_RET', 'VL_ISS', '']
        # idx = 1 (onde está 'A100')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'IND_OPER': partes[idx + 1] if idx + 1 < len(partes) else '',
            'IND_EMIT': partes[idx + 2] if idx + 2 < len(partes) else '',
            'COD_PART': partes[idx + 3] if idx + 3 < len(partes) else '',
            'COD_SIT': partes[idx + 4] if idx + 4 < len(partes) else '',
            'SER': partes[idx + 5] if idx + 5 < len(partes) else '',
            'SUB': partes[idx + 6] if idx + 6 < len(partes) else '',
            'NUM_DOC': partes[idx + 7] if idx + 7 < len(partes) else '',
            'CHV_NFE': partes[idx + 8] if idx + 8 < len(partes) else '',
            'DT_DOC': partes[idx + 9] if idx + 9 < len(partes) else '',
            'DT_EXE_SERV': partes[idx + 10] if idx + 10 < len(partes) else '',
            'VL_DOC': partes[idx + 11] if idx + 11 < len(partes) else '',
            'IND_PGTO': partes[idx + 12] if idx + 12 < len(partes) else '',
            'VL_DESC': partes[idx + 13] if idx + 13 < len(partes) else '',
            'VL_BC_PIS': partes[idx + 14] if idx + 14 < len(partes) else '',
            'VL_PIS': partes[idx + 15] if idx + 15 < len(partes) else '',
            'VL_BC_COFINS': partes[idx + 16] if idx + 16 < len(partes) else '',
            'VL_COFINS': partes[idx + 17] if idx + 17 < len(partes) else '',
            'VL_PIS_RET': partes[idx + 18] if idx + 18 < len(partes) else '',
            'VL_COFINS_RET': partes[idx + 19] if idx + 19 < len(partes) else '',
            'VL_ISS': partes[idx + 20] if idx + 20 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def obter_descricao_ind_oper(valor):
    """
    Retorna a descrição do indicador de operação.
    
    Args:
        valor: Valor do campo IND_OPER
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Entrada',
        '1': '1 - Saída'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def obter_descricao_ind_emit(valor):
    """
    Retorna a descrição do indicador do emitente do documento fiscal.
    
    Args:
        valor: Valor do campo IND_EMIT
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Emissão própria',
        '1': '1 - Terceiros'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def obter_descricao_ind_pgto(valor):
    """
    Retorna a descrição do indicador de pagamento.
    
    Args:
        valor: Valor do campo IND_PGTO
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - À vista',
        '1': '1 - A prazo',
        '2': '2 - Sem indicação'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def ler_arquivo_A100(caminho_arquivo):
    """
    Lê o arquivo A100.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A100.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A100(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A100(registros, nome_arquivo="registro_A100.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A100 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A100 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A100"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'IND_OPER',
        'IND_EMIT',
        'COD_PART',
        'COD_SIT',
        'SER',
        'SUB',
        'NUM_DOC',
        'CHV_NFE',
        'DT_DOC',
        'DT_EXE_SERV',
        'VL_DOC',
        'IND_PGTO',
        'VL_DESC',
        'VL_BC_PIS',
        'VL_PIS',
        'VL_BC_COFINS',
        'VL_COFINS',
        'VL_PIS_RET',
        'VL_COFINS_RET',
        'VL_ISS'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        dt_doc_formatada = formatar_data(registro.get('DT_DOC', ''))
        dt_exe_serv_formatada = formatar_data(registro.get('DT_EXE_SERV', ''))
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            obter_descricao_ind_oper(registro.get('IND_OPER', '')),
            obter_descricao_ind_emit(registro.get('IND_EMIT', '')),
            registro.get('COD_PART', ''),
            registro.get('COD_SIT', ''),
            registro.get('SER', '') if registro.get('SER') else '',
            registro.get('SUB', '') if registro.get('SUB') else '',
            registro.get('NUM_DOC', ''),
            registro.get('CHV_NFE', '') if registro.get('CHV_NFE') else '',
            dt_doc_formatada,
            dt_exe_serv_formatada,
            formatar_valor_monetario(registro.get('VL_DOC', '')),
            obter_descricao_ind_pgto(registro.get('IND_PGTO', '')),
            formatar_valor_monetario(registro.get('VL_DESC', '')) if registro.get('VL_DESC') else '',
            formatar_valor_monetario(registro.get('VL_BC_PIS', '')),
            formatar_valor_monetario(registro.get('VL_PIS', '')),
            formatar_valor_monetario(registro.get('VL_BC_COFINS', '')),
            formatar_valor_monetario(registro.get('VL_COFINS', '')),
            formatar_valor_monetario(registro.get('VL_PIS_RET', '')) if registro.get('VL_PIS_RET') else '',
            formatar_valor_monetario(registro.get('VL_COFINS_RET', '')) if registro.get('VL_COFINS_RET') else '',
            formatar_valor_monetario(registro.get('VL_ISS', '')) if registro.get('VL_ISS') else '',
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_A110(linha):
    """
    Parseia uma linha do registro A110 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A110 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A110' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A110'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A110':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A110
        # Estrutura: |A110|COD_INF|TXT_COMPL|
        # Após split por '|': ['', 'A110', 'COD_INF', 'TXT_COMPL', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'COD_INF': partes[idx + 1] if idx + 1 < len(partes) else '',
            'TXT_COMPL': partes[idx + 2] if idx + 2 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_A110(caminho_arquivo):
    """
    Lê o arquivo A110.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A110.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A110(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A110(registros, nome_arquivo="registro_A110.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A110 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A110 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A110"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'COD_INF',
        'TXT_COMPL'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('COD_INF', ''),
            registro.get('TXT_COMPL', '') if registro.get('TXT_COMPL') else '',
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_A170(linha):
    """
    Parseia uma linha do registro A170 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A170 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A170' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A170'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A170':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A170
        # Estrutura: |A170|NUM_ITEM|COD_ITEM|DESCR_COMPL|VL_ITEM|VL_DESC|NAT_BC_CRED|IND_ORIG_CRED|CST_PIS|VL_BC_PIS|ALIQ_PIS|VL_PIS|CST_COFINS|VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|COD_CTA|
        # Após split por '|': ['', 'A170', 'NUM_ITEM', 'COD_ITEM', 'DESCR_COMPL', 'VL_ITEM', 'VL_DESC', 'NAT_BC_CRED', 'IND_ORIG_CRED', 'CST_PIS', 'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS', 'CST_COFINS', 'VL_BC_COFINS', 'ALIQ_COFINS', 'VL_COFINS', 'COD_CTA', '']
        # idx = 1 (onde está 'A170')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'NUM_ITEM': partes[idx + 1] if idx + 1 < len(partes) else '',
            'COD_ITEM': partes[idx + 2] if idx + 2 < len(partes) else '',
            'DESCR_COMPL': partes[idx + 3] if idx + 3 < len(partes) else '',
            'VL_ITEM': partes[idx + 4] if idx + 4 < len(partes) else '',
            'VL_DESC': partes[idx + 5] if idx + 5 < len(partes) else '',
            'NAT_BC_CRED': partes[idx + 6] if idx + 6 < len(partes) else '',
            'IND_ORIG_CRED': partes[idx + 7] if idx + 7 < len(partes) else '',
            'CST_PIS': partes[idx + 8] if idx + 8 < len(partes) else '',
            'VL_BC_PIS': partes[idx + 9] if idx + 9 < len(partes) else '',
            'ALIQ_PIS': partes[idx + 10] if idx + 10 < len(partes) else '',
            'VL_PIS': partes[idx + 11] if idx + 11 < len(partes) else '',
            'CST_COFINS': partes[idx + 12] if idx + 12 < len(partes) else '',
            'VL_BC_COFINS': partes[idx + 13] if idx + 13 < len(partes) else '',
            'ALIQ_COFINS': partes[idx + 14] if idx + 14 < len(partes) else '',
            'VL_COFINS': partes[idx + 15] if idx + 15 < len(partes) else '',
            'COD_CTA': partes[idx + 16] if idx + 16 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def obter_descricao_ind_orig_cred(valor):
    """
    Retorna a descrição do indicador da origem do crédito.
    
    Args:
        valor: Valor do campo IND_ORIG_CRED
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Operação no Mercado Interno',
        '1': '1 - Operação de Importação',
        '2': '2 - Operação no Mercado Interno - Aquisição de Serviços',
        '3': '3 - Operação no Mercado Interno - Aquisição de Bens',
        '4': '4 - Operação no Mercado Interno - Aquisição de Bens e Serviços',
        '5': '5 - Operação no Mercado Interno - Aquisição de Bens para Revenda',
        '6': '6 - Operação no Mercado Interno - Aquisição de Bens para Industrialização',
        '7': '7 - Operação no Mercado Interno - Aquisição de Bens para Uso ou Consumo',
        '8': '8 - Operação no Mercado Interno - Aquisição de Bens para Ativo Fixo',
        '9': '9 - Operação no Mercado Interno - Aquisição de Serviços para Uso ou Consumo'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def ler_arquivo_A170(caminho_arquivo):
    """
    Lê o arquivo A170.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A170.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A170(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A170(registros, nome_arquivo="registro_A170.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A170 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A170 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A170"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'NUM_ITEM',
        'COD_ITEM',
        'DESCR_COMPL',
        'VL_ITEM',
        'VL_DESC',
        'NAT_BC_CRED',
        'IND_ORIG_CRED',
        'CST_PIS',
        'VL_BC_PIS',
        'ALIQ_PIS',
        'VL_PIS',
        'CST_COFINS',
        'VL_BC_COFINS',
        'ALIQ_COFINS',
        'VL_COFINS',
        'COD_CTA'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('NUM_ITEM', ''),
            registro.get('COD_ITEM', ''),
            registro.get('DESCR_COMPL', ''),
            formatar_valor_monetario(registro.get('VL_ITEM', '')),
            formatar_valor_monetario(registro.get('VL_DESC', '')) if registro.get('VL_DESC') else '',
            registro.get('NAT_BC_CRED', ''),
            obter_descricao_ind_orig_cred(registro.get('IND_ORIG_CRED', '')),
            registro.get('CST_PIS', ''),
            formatar_valor_monetario(registro.get('VL_BC_PIS', '')),
            formatar_valor_monetario(registro.get('ALIQ_PIS', '')),
            formatar_valor_monetario(registro.get('VL_PIS', '')),
            registro.get('CST_COFINS', ''),
            formatar_valor_monetario(registro.get('VL_BC_COFINS', '')),
            formatar_valor_monetario(registro.get('ALIQ_COFINS', '')),
            formatar_valor_monetario(registro.get('VL_COFINS', '')),
            registro.get('COD_CTA', '') if registro.get('COD_CTA') else '',
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_A990(linha):
    """
    Parseia uma linha do registro A990 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro A990 ou None se inválido
    """
    linha = linha.strip()
    if not linha or 'A990' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do 'A990'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == 'A990':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro A990
        # Estrutura: |A990|QTD_LIN_A|
        # Após split por '|': ['', 'A990', 'QTD_LIN_A', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'QTD_LIN_A': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_A990(caminho_arquivo):
    """
    Lê o arquivo A990.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo A990.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_A990(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_A990(registros, nome_arquivo="registro_A990.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros A990 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro A990 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro A990"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'QTD_LIN_A'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('QTD_LIN_A', ''),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    # Caminho para o arquivo A001.txt
    caminho_arquivo_A001 = "blocos_separados/bloco_A/A001.txt"
    
    print("Lendo arquivo A001.txt do bloco A...")
    registros_A001 = ler_arquivo_A001(caminho_arquivo_A001)
    
    # Gera nome do arquivo Excel com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo_excel_A001 = f"registro_A001_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A001(registros_A001, nome_arquivo_excel_A001)
    
    # Caminho para o arquivo A010.txt
    caminho_arquivo_A010 = "blocos_separados/bloco_A/A010.txt"
    
    print("\nLendo arquivo A010.txt do bloco A...")
    registros_A010 = ler_arquivo_A010(caminho_arquivo_A010)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_A010 = f"registro_A010_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A010(registros_A010, nome_arquivo_excel_A010)
    
    # Caminho para o arquivo A100.txt
    caminho_arquivo_A100 = "blocos_separados/bloco_A/A100.txt"
    
    print("\nLendo arquivo A100.txt do bloco A...")
    registros_A100 = ler_arquivo_A100(caminho_arquivo_A100)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_A100 = f"registro_A100_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A100(registros_A100, nome_arquivo_excel_A100)
    
    # Caminho para o arquivo A110.txt
    caminho_arquivo_A110 = "blocos_separados/bloco_A/A110.txt"
    
    print("\nLendo arquivo A110.txt do bloco A...")
    registros_A110 = ler_arquivo_A110(caminho_arquivo_A110)

    nome_arquivo_excel_A110 = f"registro_A110_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A110(registros_A110, nome_arquivo_excel_A110)

    caminho_arquivo_A170 = "blocos_separados/bloco_A/A170.txt"
    
    print("\nLendo arquivo A170.txt do bloco A...")
    registros_A170 = ler_arquivo_A170(caminho_arquivo_A170)
    
    nome_arquivo_excel_A170 = f"registro_A170_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A170(registros_A170, nome_arquivo_excel_A170)

    caminho_arquivo_A990 = "blocos_separados/bloco_A/A990.txt"
    
    print("\nLendo arquivo A990.txt do bloco A...")
    registros_A990 = ler_arquivo_A990(caminho_arquivo_A990)

    nome_arquivo_excel_A990 = f"registro_A990_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_A990(registros_A990, nome_arquivo_excel_A990)

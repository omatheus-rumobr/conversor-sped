import os
from datetime import datetime

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook


def parse_registro_1001(linha):
    """
    Parseia uma linha do registro 1001 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1001 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1001' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1001'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1001':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1001
        # Estrutura: |1001|IND_MOV|
        # Após split por '|': ['', '1001', 'IND_MOV', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'IND_MOV': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def obter_descricao_ind_mov(valor):
    """
    Retorna a descrição do indicador de movimento.
    
    Args:
        valor: Valor do campo IND_MOV
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Bloco com dados informados',
        '1': '1 - Bloco sem dados'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def ler_arquivo_1001(caminho_arquivo):
    """
    Lê o arquivo 1001.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1001.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1001(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1001(registros, nome_arquivo="registro_1001.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1001 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1001 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1001"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'IND_MOV'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            obter_descricao_ind_mov(registro.get('IND_MOV', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_1100(linha):
    """
    Parseia uma linha do registro 1100 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1100 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1100' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1100'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1100':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1100
        # Estrutura: |1100|PER_APU_CRED|ORIG_CRED|CNPJ_SUC|COD_CRED|VL_CRED_APU|VL_CRED_EXT_APU|VL_TOT_CRED_APU|VL_CRED_DESC_PA_ANT|VL_CRED_PER_PA_ANT|VL_CRED_DCOMP_PA_ANT|SD_CRED_DISP_EFD|VL_CRED_DESC_EFD|VL_CRED_PER_EFD|VL_CRED_DCOMP_EFD|VL_CRED_TRANS|VL_CRED_OUT|SLD_CRED_FIM|
        # Após split por '|': ['', '1100', 'PER_APU_CRED', 'ORIG_CRED', '', 'COD_CRED', 'VL_CRED_APU', ...]
        # idx = 1 (onde está '1100')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'PER_APU_CRED': partes[idx + 1] if idx + 1 < len(partes) else '',
            'ORIG_CRED': partes[idx + 2] if idx + 2 < len(partes) else '',
            'CNPJ_SUC': partes[idx + 3] if idx + 3 < len(partes) else '',
            'COD_CRED': partes[idx + 4] if idx + 4 < len(partes) else '',
            'VL_CRED_APU': partes[idx + 5] if idx + 5 < len(partes) else '',
            'VL_CRED_EXT_APU': partes[idx + 6] if idx + 6 < len(partes) else '',
            'VL_TOT_CRED_APU': partes[idx + 7] if idx + 7 < len(partes) else '',
            'VL_CRED_DESC_PA_ANT': partes[idx + 8] if idx + 8 < len(partes) else '',
            'VL_CRED_PER_PA_ANT': partes[idx + 9] if idx + 9 < len(partes) else '',
            'VL_CRED_DCOMP_PA_ANT': partes[idx + 10] if idx + 10 < len(partes) else '',
            'SD_CRED_DISP_EFD': partes[idx + 11] if idx + 11 < len(partes) else '',
            'VL_CRED_DESC_EFD': partes[idx + 12] if idx + 12 < len(partes) else '',
            'VL_CRED_PER_EFD': partes[idx + 13] if idx + 13 < len(partes) else '',
            'VL_CRED_DCOMP_EFD': partes[idx + 14] if idx + 14 < len(partes) else '',
            'VL_CRED_TRANS': partes[idx + 15] if idx + 15 < len(partes) else '',
            'VL_CRED_OUT': partes[idx + 16] if idx + 16 < len(partes) else '',
            'SLD_CRED_FIM': partes[idx + 17] if idx + 17 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def formatar_periodo(periodo_str):
    """
    Formata período do formato MMYYYY para MM/YYYY.
    
    Args:
        periodo_str: String com período no formato MMYYYY
        
    Returns:
        String formatada como MM/YYYY ou a string original se inválida
    """
    if not periodo_str or len(periodo_str) != 6:
        return periodo_str
    
    try:
        return f"{periodo_str[0:2]}/{periodo_str[2:6]}"
    except:
        return periodo_str


def obter_descricao_orig_cred(valor):
    """
    Retorna a descrição da origem do crédito.
    
    Args:
        valor: Valor do campo ORIG_CRED
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '01': '01 - Crédito próprio',
        '02': '02 - Crédito transferido por pessoa jurídica sucedida'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def formatar_valor_monetario(valor_str):
    """
    Formata valor monetário substituindo vírgula por ponto para compatibilidade com Excel.
    
    Args:
        valor_str: String com valor monetário (formato brasileiro com vírgula)
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not valor_str:
        return valor_str
    
    try:
        # Substitui vírgula por ponto para compatibilidade com Excel
        return valor_str.replace(',', '.')
    except:
        return valor_str


def formatar_cnpj(cnpj_str):
    """
    Formata CNPJ para o formato XX.XXX.XXX/XXXX-XX.
    
    Args:
        cnpj_str: String com CNPJ sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cnpj_str or len(cnpj_str) != 14:
        return cnpj_str
    
    try:
        return f"{cnpj_str[0:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:14]}"
    except:
        return cnpj_str


def ler_arquivo_1100(caminho_arquivo):
    """
    Lê o arquivo 1100.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1100.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1100(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1100(registros, nome_arquivo="registro_1100.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1100 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1100 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1100"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'PER_APU_CRED',
        'ORIG_CRED',
        'CNPJ_SUC',
        'COD_CRED',
        'VL_CRED_APU',
        'VL_CRED_EXT_APU',
        'VL_TOT_CRED_APU',
        'VL_CRED_DESC_PA_ANT',
        'VL_CRED_PER_PA_ANT',
        'VL_CRED_DCOMP_PA_ANT',
        'SD_CRED_DISP_EFD',
        'VL_CRED_DESC_EFD',
        'VL_CRED_PER_EFD',
        'VL_CRED_DCOMP_EFD',
        'VL_CRED_TRANS',
        'VL_CRED_OUT',
        'SLD_CRED_FIM'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        periodo_formatado = formatar_periodo(registro.get('PER_APU_CRED', ''))
        cnpj_formatado = formatar_cnpj(registro.get('CNPJ_SUC', '')) if registro.get('CNPJ_SUC') else ''
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            periodo_formatado,
            obter_descricao_orig_cred(registro.get('ORIG_CRED', '')),
            cnpj_formatado,
            registro.get('COD_CRED', ''),
            formatar_valor_monetario(registro.get('VL_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_EXT_APU', '')),
            formatar_valor_monetario(registro.get('VL_TOT_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DCOMP_PA_ANT', '')),
            formatar_valor_monetario(registro.get('SD_CRED_DISP_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DCOMP_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_TRANS', '')),
            formatar_valor_monetario(registro.get('VL_CRED_OUT', '')),
            formatar_valor_monetario(registro.get('SLD_CRED_FIM', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_1300(linha):
    """
    Parseia uma linha do registro 1300 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1300 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1300' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1300'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1300':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1300
        # Estrutura: |1300|COD_ITEM|PER_APU_CRED|VL_CRED_APU|VL_CRED_EXT_APU|VL_CRED_DESC_PA_ANT|VL_CRED_PER_PA_ANT|VL_TOT_CRED_APU|
        # Após split por '|': ['', '1300', 'COD_ITEM', 'PER_APU_CRED', 'VL_CRED_APU', 'VL_CRED_EXT_APU', 'VL_CRED_DESC_PA_ANT', 'VL_CRED_PER_PA_ANT', 'VL_TOT_CRED_APU', '']
        # idx = 1 (onde está '1300')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'COD_ITEM': partes[idx + 1] if idx + 1 < len(partes) else '',
            'PER_APU_CRED': partes[idx + 2] if idx + 2 < len(partes) else '',
            'VL_CRED_APU': partes[idx + 3] if idx + 3 < len(partes) else '',
            'VL_CRED_EXT_APU': partes[idx + 4] if idx + 4 < len(partes) else '',
            'VL_CRED_DESC_PA_ANT': partes[idx + 5] if idx + 5 < len(partes) else '',
            'VL_CRED_PER_PA_ANT': partes[idx + 6] if idx + 6 < len(partes) else '',
            'VL_TOT_CRED_APU': partes[idx + 7] if idx + 7 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def formatar_periodo_1300(periodo_str):
    """
    Formata período do formato MMMMAA para MM/AAAA.
    
    Args:
        periodo_str: String com período no formato MMMMAA
        
    Returns:
        String formatada como MM/AAAA ou a string original se inválida
    """
    if not periodo_str or len(periodo_str) != 6:
        return periodo_str
    
    try:
        # Formato MMMMAA -> MM/AAAA
        return f"{periodo_str[0:2]}/{periodo_str[2:6]}"
    except:
        return periodo_str


def ler_arquivo_1300(caminho_arquivo):
    """
    Lê o arquivo 1300.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1300.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1300(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1300(registros, nome_arquivo="registro_1300.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1300 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1300 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1300"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'COD_ITEM',
        'PER_APU_CRED',
        'VL_CRED_APU',
        'VL_CRED_EXT_APU',
        'VL_CRED_DESC_PA_ANT',
        'VL_CRED_PER_PA_ANT',
        'VL_TOT_CRED_APU'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        periodo_formatado = formatar_periodo_1300(registro.get('PER_APU_CRED', ''))
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('COD_ITEM', ''),
            periodo_formatado,
            formatar_valor_monetario(registro.get('VL_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_EXT_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_TOT_CRED_APU', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_1500(linha):
    """
    Parseia uma linha do registro 1500 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1500 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1500' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1500'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1500':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1500
        # Estrutura: |1500|PER_APU_CRED|ORIG_CRED|CNPJ_SUC|COD_CRED|VL_CRED_APU|VL_CRED_EXT_APU|VL_TOT_CRED_APU|VL_CRED_DESC_PA_ANT|VL_CRED_PER_PA_ANT|VL_CRED_DCOMP_PA_ANT|SD_CRED_DISP_EFD|VL_CRED_DESC_EFD|VL_CRED_PER_EFD|VL_CRED_DCOMP_EFD|VL_CRED_TRANS|VL_CRED_OUT|SLD_CRED_FIM|
        # Após split por '|': ['', '1500', 'PER_APU_CRED', 'ORIG_CRED', '', 'COD_CRED', 'VL_CRED_APU', ...]
        # idx = 1 (onde está '1500')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'PER_APU_CRED': partes[idx + 1] if idx + 1 < len(partes) else '',
            'ORIG_CRED': partes[idx + 2] if idx + 2 < len(partes) else '',
            'CNPJ_SUC': partes[idx + 3] if idx + 3 < len(partes) else '',
            'COD_CRED': partes[idx + 4] if idx + 4 < len(partes) else '',
            'VL_CRED_APU': partes[idx + 5] if idx + 5 < len(partes) else '',
            'VL_CRED_EXT_APU': partes[idx + 6] if idx + 6 < len(partes) else '',
            'VL_TOT_CRED_APU': partes[idx + 7] if idx + 7 < len(partes) else '',
            'VL_CRED_DESC_PA_ANT': partes[idx + 8] if idx + 8 < len(partes) else '',
            'VL_CRED_PER_PA_ANT': partes[idx + 9] if idx + 9 < len(partes) else '',
            'VL_CRED_DCOMP_PA_ANT': partes[idx + 10] if idx + 10 < len(partes) else '',
            'SD_CRED_DISP_EFD': partes[idx + 11] if idx + 11 < len(partes) else '',
            'VL_CRED_DESC_EFD': partes[idx + 12] if idx + 12 < len(partes) else '',
            'VL_CRED_PER_EFD': partes[idx + 13] if idx + 13 < len(partes) else '',
            'VL_CRED_DCOMP_EFD': partes[idx + 14] if idx + 14 < len(partes) else '',
            'VL_CRED_TRANS': partes[idx + 15] if idx + 15 < len(partes) else '',
            'VL_CRED_OUT': partes[idx + 16] if idx + 16 < len(partes) else '',
            'SLD_CRED_FIM': partes[idx + 17] if idx + 17 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_1500(caminho_arquivo):
    """
    Lê o arquivo 1500.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1500.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1500(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1500(registros, nome_arquivo="registro_1500.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1500 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1500 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1500"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'PER_APU_CRED',
        'ORIG_CRED',
        'CNPJ_SUC',
        'COD_CRED',
        'VL_CRED_APU',
        'VL_CRED_EXT_APU',
        'VL_TOT_CRED_APU',
        'VL_CRED_DESC_PA_ANT',
        'VL_CRED_PER_PA_ANT',
        'VL_CRED_DCOMP_PA_ANT',
        'SD_CRED_DISP_EFD',
        'VL_CRED_DESC_EFD',
        'VL_CRED_PER_EFD',
        'VL_CRED_DCOMP_EFD',
        'VL_CRED_TRANS',
        'VL_CRED_OUT',
        'SLD_CRED_FIM'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        periodo_formatado = formatar_periodo(registro.get('PER_APU_CRED', ''))
        cnpj_formatado = formatar_cnpj(registro.get('CNPJ_SUC', '')) if registro.get('CNPJ_SUC') else ''
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            periodo_formatado,
            obter_descricao_orig_cred(registro.get('ORIG_CRED', '')),
            cnpj_formatado,
            registro.get('COD_CRED', ''),
            formatar_valor_monetario(registro.get('VL_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_EXT_APU', '')),
            formatar_valor_monetario(registro.get('VL_TOT_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DCOMP_PA_ANT', '')),
            formatar_valor_monetario(registro.get('SD_CRED_DISP_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DCOMP_EFD', '')),
            formatar_valor_monetario(registro.get('VL_CRED_TRANS', '')),
            formatar_valor_monetario(registro.get('VL_CRED_OUT', '')),
            formatar_valor_monetario(registro.get('SLD_CRED_FIM', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_1700(linha):
    """
    Parseia uma linha do registro 1700 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1700 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1700' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1700'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1700':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1700
        # Estrutura: |1700|COD_ITEM|PER_APU_CRED|VL_CRED_APU|VL_CRED_EXT_APU|VL_CRED_DESC_PA_ANT|VL_CRED_PER_PA_ANT|VL_TOT_CRED_APU|
        # Após split por '|': ['', '1700', 'COD_ITEM', 'PER_APU_CRED', 'VL_CRED_APU', 'VL_CRED_EXT_APU', 'VL_CRED_DESC_PA_ANT', 'VL_CRED_PER_PA_ANT', 'VL_TOT_CRED_APU', '']
        # idx = 1 (onde está '1700')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'COD_ITEM': partes[idx + 1] if idx + 1 < len(partes) else '',
            'PER_APU_CRED': partes[idx + 2] if idx + 2 < len(partes) else '',
            'VL_CRED_APU': partes[idx + 3] if idx + 3 < len(partes) else '',
            'VL_CRED_EXT_APU': partes[idx + 4] if idx + 4 < len(partes) else '',
            'VL_CRED_DESC_PA_ANT': partes[idx + 5] if idx + 5 < len(partes) else '',
            'VL_CRED_PER_PA_ANT': partes[idx + 6] if idx + 6 < len(partes) else '',
            'VL_TOT_CRED_APU': partes[idx + 7] if idx + 7 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_1700(caminho_arquivo):
    """
    Lê o arquivo 1700.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1700.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1700(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1700(registros, nome_arquivo="registro_1700.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1700 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1700 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1700"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'COD_ITEM',
        'PER_APU_CRED',
        'VL_CRED_APU',
        'VL_CRED_EXT_APU',
        'VL_CRED_DESC_PA_ANT',
        'VL_CRED_PER_PA_ANT',
        'VL_TOT_CRED_APU'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Formata os dados
        periodo_formatado = formatar_periodo_1300(registro.get('PER_APU_CRED', ''))
        
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('COD_ITEM', ''),
            periodo_formatado,
            formatar_valor_monetario(registro.get('VL_CRED_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_EXT_APU', '')),
            formatar_valor_monetario(registro.get('VL_CRED_DESC_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_CRED_PER_PA_ANT', '')),
            formatar_valor_monetario(registro.get('VL_TOT_CRED_APU', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_1990(linha):
    """
    Parseia uma linha do registro 1990 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 1990 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '1990' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '1990'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '1990':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 1990
        # Estrutura: |1990|QTD_LIN_1|
        # Após split por '|': ['', '1990', 'QTD_LIN_1', '']
        # idx = 1 (onde está '1990')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'QTD_LIN_1': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_1990(caminho_arquivo):
    """
    Lê o arquivo 1990.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 1990.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_1990(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_1990(registros, nome_arquivo="registro_1990.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 1990 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 1990 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 1990"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'QTD_LIN_1'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('QTD_LIN_1', ''),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    # Caminho para o arquivo 1001.txt
    caminho_arquivo_1001 = "blocos_separados/bloco_1/1001.txt"
    
    print("Lendo arquivo 1001.txt do bloco 1...")
    registros_1001 = ler_arquivo_1001(caminho_arquivo_1001)
    
    # Gera nome do arquivo Excel com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo_excel_1001 = f"registro_1001_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1001(registros_1001, nome_arquivo_excel_1001)
    
    # Caminho para o arquivo 1100.txt
    caminho_arquivo_1100 = "blocos_separados/bloco_1/1100.txt"
    
    print("\nLendo arquivo 1100.txt do bloco 1...")
    registros_1100 = ler_arquivo_1100(caminho_arquivo_1100)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_1100 = f"registro_1100_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1100(registros_1100, nome_arquivo_excel_1100)
    
    # Caminho para o arquivo 1300.txt
    caminho_arquivo_1300 = "blocos_separados/bloco_1/1300.txt"
    
    print("\nLendo arquivo 1300.txt do bloco 1...")
    registros_1300 = ler_arquivo_1300(caminho_arquivo_1300)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_1300 = f"registro_1300_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1300(registros_1300, nome_arquivo_excel_1300)
    
    # Caminho para o arquivo 1500.txt
    caminho_arquivo_1500 = "blocos_separados/bloco_1/1500.txt"
    
    print("\nLendo arquivo 1500.txt do bloco 1...")
    registros_1500 = ler_arquivo_1500(caminho_arquivo_1500)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_1500 = f"registro_1500_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1500(registros_1500, nome_arquivo_excel_1500)
    
    # Caminho para o arquivo 1700.txt
    caminho_arquivo_1700 = "blocos_separados/bloco_1/1700.txt"
    
    print("\nLendo arquivo 1700.txt do bloco 1...")
    registros_1700 = ler_arquivo_1700(caminho_arquivo_1700)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_1700 = f"registro_1700_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1700(registros_1700, nome_arquivo_excel_1700)
    
    # Caminho para o arquivo 1990.txt
    caminho_arquivo_1990 = "blocos_separados/bloco_1/1990.txt"
    
    print("\nLendo arquivo 1990.txt do bloco 1...")
    registros_1990 = ler_arquivo_1990(caminho_arquivo_1990)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_1990 = f"registro_1990_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_1990(registros_1990, nome_arquivo_excel_1990)
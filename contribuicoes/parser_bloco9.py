import os
from datetime import datetime

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook


def parse_registro_9001(linha):
    """
    Parseia uma linha do registro 9001 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 9001 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '9001' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '9001'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '9001':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 9001
        # Estrutura: |9001|IND_MOV|
        # Após split por '|': ['', '9001', 'IND_MOV', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'IND_MOV': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def obter_descricao_ind_mov(valor):
    """
    Retorna a descrição do indicador de movimento.
    
    Args:
        valor: Valor do campo IND_MOV
        
    Returns:
        String com descrição ou valor original
    """
    descricoes = {
        '0': '0 - Bloco com dados informados',
        '1': '1 - Bloco sem dados'
    }
    
    if valor in descricoes:
        return descricoes[valor]
    
    return valor


def ler_arquivo_9001(caminho_arquivo):
    """
    Lê o arquivo 9001.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 9001.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_9001(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_9001(registros, nome_arquivo="registro_9001.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 9001 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 9001 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 9001"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'IND_MOV'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            obter_descricao_ind_mov(registro.get('IND_MOV', '')),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_9900(linha):
    """
    Parseia uma linha do registro 9900 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 9900 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '9900' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '9900'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '9900':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 9900
        # Estrutura: |9900|REG_BLC|QTD_REG_BLC|
        # Após split por '|': ['', '9900', 'REG_BLC', 'QTD_REG_BLC', '']
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'REG_BLC': partes[idx + 1] if idx + 1 < len(partes) else '',
            'QTD_REG_BLC': partes[idx + 2] if idx + 2 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_9900(caminho_arquivo):
    """
    Lê o arquivo 9900.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 9900.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_9900(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_9900(registros, nome_arquivo="registro_9900.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 9900 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 9900 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 9900"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'REG_BLC',
        'QTD_REG_BLC'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('REG_BLC', ''),
            registro.get('QTD_REG_BLC', ''),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def parse_registro_9990(linha):
    """
    Parseia uma linha do registro 9990 do SPED EFD Contribuições.
    
    Args:
        linha: String com a linha do arquivo SPED
        
    Returns:
        Dicionário com os campos do registro 9990 ou None se inválido
    """
    linha = linha.strip()
    if not linha or '9990' not in linha:
        return None
    
    # Divide a linha por pipes
    partes = linha.split('|')
    
    # Procura o índice do '9990'
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == '9990':
                idx = i
                break
        
        if idx is None:
            return None
        
        # Extrai os campos conforme o layout do registro 9990
        # Estrutura: |9990|QTD_LIN_9|
        # Após split por '|': ['', '9990', 'QTD_LIN_9', '']
        # idx = 1 (onde está '9990')
        campos = {
            'REG': partes[idx] if idx < len(partes) else '',
            'QTD_LIN_9': partes[idx + 1] if idx + 1 < len(partes) else '',
        }
        
        return campos
        
    except (IndexError, ValueError) as e:
        print(f"Erro ao parsear linha: {e}")
        return None


def ler_arquivo_9990(caminho_arquivo):
    """
    Lê o arquivo 9990.txt e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo 9990.txt
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Erro: Arquivo '{caminho_arquivo}' não encontrado!")
        return registros
    
    # Tenta diferentes encodings
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_9990(linha)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro_9990(registros, nome_arquivo="registro_9990.xlsx"):
    """
    Cria um arquivo Excel (.xlsx) com os registros 9990 sem formatação.
    
    Args:
        registros: Lista de dicionários com os registros
        nome_arquivo: Nome do arquivo Excel a ser criado
    """
    if not registros:
        print("Nenhum registro 9990 encontrado!")
        return None
    
    # Cria um workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro 9990"
    
    # Define os cabeçalhos das colunas
    cabecalhos = [
        'REG',
        'QTD_LIN_9'
    ]
    
    # Escreve os cabeçalhos na primeira linha
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    # Preenche os dados
    linha_atual = 2
    
    for registro in registros:
        # Prepara os valores na ordem dos cabeçalhos
        valores = [
            registro.get('REG', ''),
            registro.get('QTD_LIN_9', ''),
        ]
        
        # Escreve os valores na linha (sem formatação)
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor if valor else '')
        
        linha_atual += 1
    
    # Salva o arquivo
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    # Caminho para o arquivo 9001.txt
    caminho_arquivo_9001 = "blocos_separados/bloco_9/9001.txt"
    
    print("Lendo arquivo 9001.txt do bloco 9...")
    registros_9001 = ler_arquivo_9001(caminho_arquivo_9001)
    
    # Gera nome do arquivo Excel com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo_excel_9001 = f"registro_9001_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_9001(registros_9001, nome_arquivo_excel_9001)
    
    # Caminho para o arquivo 9900.txt
    caminho_arquivo_9900 = "blocos_separados/bloco_9/9900.txt"
    
    print("\nLendo arquivo 9900.txt do bloco 9...")
    registros_9900 = ler_arquivo_9900(caminho_arquivo_9900)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_9900 = f"registro_9900_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_9900(registros_9900, nome_arquivo_excel_9900)
    
    # Caminho para o arquivo 9990.txt
    caminho_arquivo_9990 = "blocos_separados/bloco_9/9990.txt"
    
    print("\nLendo arquivo 9990.txt do bloco 9...")
    registros_9990 = ler_arquivo_9990(caminho_arquivo_9990)
    
    # Gera nome do arquivo Excel com timestamp
    nome_arquivo_excel_9990 = f"registro_9990_{timestamp}.xlsx"
    
    criar_arquivo_excel_registro_9990(registros_9990, nome_arquivo_excel_9990)
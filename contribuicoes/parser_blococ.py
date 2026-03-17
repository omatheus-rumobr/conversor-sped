import os
import re
from datetime import datetime
from collections import OrderedDict

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook


def ler_definicoes_registros(arquivo_markdown):
    """
    Lê o arquivo markdown e extrai as definições de todos os registros.
    
    Args:
        arquivo_markdown: Caminho para o arquivo sped_bloco_c.md
        
    Returns:
        Dicionário com código do registro como chave e lista de campos como valor
        Exemplo: {'C001': [{'numero': '01', 'campo': 'REG', 'tipo': 'C', ...}, ...]}
    """
    definicoes = OrderedDict()
    
    if not os.path.exists(arquivo_markdown):
        print(f"Erro: Arquivo '{arquivo_markdown}' não encontrado!")
        return definicoes
    
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_markdown, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding do markdown detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo markdown!")
        return definicoes
    
    try:
        conteudo = arquivo.read()
    finally:
        arquivo.close()
    
    padrao_registro = r'Registro\s+(C\d+):'
    
    matches = list(re.finditer(padrao_registro, conteudo, re.IGNORECASE))
    
    for i, match in enumerate(matches):
        codigo_registro = match.group(1).upper()
        inicio = match.end()
        
        if i + 1 < len(matches):
            fim = matches[i + 1].start()
        else:
            fim = len(conteudo)
        
        secao = conteudo[inicio:fim]
        
        padroes_tabela = [
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[çc][ãa]o\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descricao\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[çc]ao\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
            r'\|\s*N[°ºoO]\s*\|\s*Campo\s*\|\s*Descri[ãa]o\s*\|\s*Tipo\s*\|\s*Tam\s*\|\s*Dec\s*\|\s*Obrig\s*\|',
        ]
        
        match_tabela = None
        for padrao in padroes_tabela:
            match_tabela = re.search(padrao, secao, re.IGNORECASE)
            if match_tabela:
                break
        
        if match_tabela:
            inicio_tabela = match_tabela.end()
            linhas = secao[inicio_tabela:].split('\n')
            
            campos = []
            encontrou_separador = False
            
            for linha in linhas:
                linha = linha.strip()
                
                if linha.startswith('|') and '----' in linha:
                    encontrou_separador = True
                    continue
                
                if (linha.startswith('Observações') or linha.startswith('Observacoes') or 
                    linha.startswith('Nível') or linha.startswith('Nivel')):
                    if campos:
                        break
                    continue
                
                if not encontrou_separador:
                    continue
                
                if linha.startswith('|') and linha.count('|') >= 3:
                    partes = [p.strip() for p in linha.split('|')]
                    
                    if partes and not partes[0]:
                        partes = partes[1:]

                    if partes and not partes[-1]:
                        partes = partes[:-1]
                    
                    if len(partes) >= 7:
                        try:
                            if partes[0].isdigit():
                                campo = {
                                    'numero': partes[0],
                                    'campo': partes[1],
                                    'descricao': partes[2],
                                    'tipo': partes[3],
                                    'tamanho': partes[4],
                                    'decimais': partes[5],
                                    'obrigatorio': partes[6]
                                }
                                campos.append(campo)
                        except (IndexError, ValueError):
                            continue
            
            if campos:
                definicoes[codigo_registro] = campos
                print(f"Registro {codigo_registro} encontrado com {len(campos)} campos")
    
    print(f"\nTotal de registros encontrados: {len(definicoes)}")
    return definicoes


def parse_registro_generico(linha, codigo_registro, campos_definicao):
    """
    Parseia uma linha do SPED de forma genérica baseado nas definições.
    
    Args:
        linha: String com a linha do arquivo SPED
        codigo_registro: Código do registro (ex: 'C001')
        campos_definicao: Lista de dicionários com as definições dos campos
        
    Returns:
        Dicionário com os campos parseados ou None se inválido
    """
    linha = linha.strip()
    if not linha or codigo_registro.upper() not in linha:
        return None
    
    partes = linha.split('|')
    
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip().upper() == codigo_registro.upper():
                idx = i
                break
        
        if idx is None:
            return None
        
        campos = {}
        for campo_def in campos_definicao:
            numero_campo = int(campo_def['numero'])
            nome_campo = campo_def['campo']
            
            posicao = idx + (numero_campo - 1)
            
            if posicao < len(partes):
                valor = partes[posicao].strip()
                campos[nome_campo] = valor
            else:
                campos[nome_campo] = ''
        
        return campos
        
    except (IndexError, ValueError, KeyError) as e:
        print(f"Erro ao parsear linha do registro {codigo_registro}: {e}")
        return None


def formatar_cnpj(cnpj_str):
    """
    Formata CNPJ para o formato XX.XXX.XXX/XXXX-XX.
    
    Args:
        cnpj_str: String com CNPJ sem formatação
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not cnpj_str or len(cnpj_str) != 14:
        return cnpj_str
    
    try:
        return f"{cnpj_str[0:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:14]}"
    except:
        return cnpj_str


def formatar_data(data_str):
    """
    Formata data do formato DDMMYYYY para DD/MM/YYYY.
    
    Args:
        data_str: String com data no formato DDMMYYYY
        
    Returns:
        String formatada como DD/MM/YYYY ou a string original se inválida
    """
    if not data_str or len(data_str) != 8:
        return data_str
    
    try:
        return f"{data_str[0:2]}/{data_str[2:4]}/{data_str[4:8]}"
    except:
        return data_str


def formatar_valor_monetario(valor_str):
    """
    Formata valor monetário substituindo vírgula por ponto para compatibilidade com Excel.
    
    Args:
        valor_str: String com valor monetário (formato brasileiro com vírgula)
        
    Returns:
        String formatada ou a string original se inválida
    """
    if not valor_str:
        return valor_str
    
    try:
        return valor_str.replace(',', '.')
    except:
        return valor_str


def obter_descricao_campo(campo, valor):
    """
    Retorna descrição formatada para campos específicos baseado no valor.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        
    Returns:
        String com descrição formatada ou valor original
    """
    if not valor:
        return valor
    
    descricoes = {
        'IND_MOV': {
            '0': '0 - Bloco com dados informados',
            '1': '1 - Bloco sem dados informados'
        },
        'IND_OPER': {
            '0': '0 - Entrada',
            '1': '1 - Saída'
        },
        'IND_EMIT': {
            '0': '0 - Emissão própria',
            '1': '1 - Terceiros'
        },
        'IND_PGTO': {
            '0': '0 - À vista',
            '1': '1 - A prazo',
            '2': '2 - Sem indicação',
            '9': '9 - Sem pagamento'
        },
        'IND_ESCRI': {
            '1': '1 - Apuração com base nos registros de consolidação',
            '2': '2 - Apuração com base no registro individualizado'
        }
    }
    
    if campo in descricoes and valor in descricoes[campo]:
        return descricoes[campo][valor]
    
    return valor


def aplicar_formatacao(campo, valor, tipo_campo):
    """
    Aplica formatação específica baseada no tipo e nome do campo.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        tipo_campo: Tipo do campo (C, N, etc)
        
    Returns:
        Valor formatado
    """
    if not valor:
        return valor
    
    if campo == 'CNPJ' and len(valor) == 14:
        return formatar_cnpj(valor)
    
    if campo.startswith('DT_') or campo.startswith('DTE_'):
        if len(valor) == 8 and valor.isdigit():
            return formatar_data(valor)
    
    if campo.startswith('VL_') and tipo_campo == 'N':
        return formatar_valor_monetario(valor)
    
    return obter_descricao_campo(campo, valor)


def ler_arquivo_registro(caminho_arquivo, codigo_registro, campos_definicao):
    """
    Lê o arquivo de um registro específico e retorna os registros parseados.
    
    Args:
        caminho_arquivo: Caminho para o arquivo do registro
        codigo_registro: Código do registro (ex: 'C001')
        campos_definicao: Lista de dicionários com as definições dos campos
        
    Returns:
        Lista de dicionários com os registros parseados
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo '{caminho_arquivo}' não encontrado (pode não existir registros deste tipo)")
        return registros
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo '{caminho_arquivo}'!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_generico(linha, codigo_registro, campos_definicao)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def ler_dados_bloco_0():
    """
    Lê os dados do bloco 0 necessários para as colunas iniciais.
    
    Returns:
        Dicionário com cnpj, ie e periodo, ou valores vazios se não encontrados
    """
    dados = {
        'cnpj': '',
        'ie': '',
        'periodo': ''
    }
    
    # Lê registro 0000 para obter CNPJ e período
    caminho_0000 = "blocos_separados/bloco_0/0000.txt"
    if os.path.exists(caminho_0000):
        encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
        arquivo = None
        
        for encoding in encodings:
            try:
                arquivo = open(caminho_0000, 'r', encoding=encoding)
                arquivo.readline()
                arquivo.seek(0)
                break
            except (UnicodeDecodeError, UnicodeError):
                if arquivo:
                    arquivo.close()
                continue
        
        if arquivo:
            try:
                for linha in arquivo:
                    linha = linha.strip()
                    if linha and linha.startswith('|'):
                        partes = linha.split('|')
                        # Verifica se é o registro 0000 (campo REG está na posição 1 após o split)
                        if len(partes) > 1 and partes[1].strip() == '0000':
                            if len(partes) >= 10:
                                # CNPJ está na posição 9 (índice 9, campo 09)
                                cnpj = partes[9].strip() if len(partes) > 9 else ''
                                # DT_INI está na posição 6 (índice 6, campo 06)
                                dt_ini = partes[6].strip() if len(partes) > 6 else ''
                                # DT_FIN está na posição 7 (índice 7, campo 07)
                                dt_fin = partes[7].strip() if len(partes) > 7 else ''
                                
                                if cnpj:
                                    dados['cnpj'] = formatar_cnpj(cnpj) if len(cnpj) == 14 else cnpj
                                
                                if dt_ini and dt_fin:
                                    # Formata período como DD/MM/YYYY a DD/MM/YYYY
                                    if len(dt_ini) == 8 and len(dt_fin) == 8:
                                        periodo_ini = formatar_data(dt_ini)
                                        periodo_fin = formatar_data(dt_fin)
                                        dados['periodo'] = f"{periodo_ini} a {periodo_fin}"
                            break
            finally:
                arquivo.close()
    
    # Lê registro 0140 para obter Inscrição Estadual (IE) do estabelecimento matriz
    caminho_0140 = "blocos_separados/bloco_0/0140.txt"
    if os.path.exists(caminho_0140):
        encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
        arquivo = None
        
        for encoding in encodings:
            try:
                arquivo = open(caminho_0140, 'r', encoding=encoding)
                arquivo.readline()
                arquivo.seek(0)
                break
            except (UnicodeDecodeError, UnicodeError):
                if arquivo:
                    arquivo.close()
                continue
        
        if arquivo:
            try:
                # Pega o primeiro registro 0140 (estabelecimento matriz)
                for linha in arquivo:
                    linha = linha.strip()
                    if linha and linha.startswith('|'):
                        partes = linha.split('|')
                        # Verifica se é o registro 0140 (campo REG está na posição 1 após o split)
                        if len(partes) > 1 and partes[1].strip() == '0140':
                            if len(partes) >= 7:
                                # IE está na posição 6 (índice 6, campo 06)
                                ie = partes[6].strip() if len(partes) > 6 else ''
                                if ie:
                                    dados['ie'] = ie
                            break
            finally:
                arquivo.close()
    
    return dados


def criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo=None, dados_bloco_0=None):
    """
    Cria um arquivo Excel (.xlsx) com os registros parseados.
    
    Args:
        registros: Lista de dicionários com os registros
        codigo_registro: Código do registro (ex: 'C001')
        campos_definicao: Lista de dicionários com as definições dos campos
        nome_arquivo: Nome do arquivo Excel a ser criado (opcional)
        dados_bloco_0: Dicionário com cnpj, ie e periodo do bloco 0 (opcional)
    """
    if not registros:
        print(f"Nenhum registro {codigo_registro} encontrado!")
        return None
    
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"registro_{codigo_registro}_{timestamp}.xlsx"
    
    if dados_bloco_0 is None:
        dados_bloco_0 = ler_dados_bloco_0()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {codigo_registro}"
    
    # Adiciona as 3 primeiras colunas: CNPJ, IE, Período
    cabecalhos = ['CNPJ', 'Inscrição Estadual', 'Período']
    # Adiciona os cabeçalhos dos campos do registro
    cabecalhos.extend([campo['campo'] for campo in campos_definicao])
    
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    linha_atual = 2
    
    for registro in registros:
        valores = []
        # Adiciona os valores do bloco 0 nas primeiras 3 colunas
        valores.append(dados_bloco_0.get('cnpj', ''))
        valores.append(dados_bloco_0.get('ie', ''))
        valores.append(dados_bloco_0.get('periodo', ''))
        
        # Adiciona os valores dos campos do registro
        for campo_def in campos_definicao:
            nome_campo = campo_def['campo']
            tipo_campo = campo_def['tipo']
            valor = registro.get(nome_campo, '')
            
            valor_formatado = aplicar_formatacao(nome_campo, valor, tipo_campo)
            valores.append(valor_formatado if valor_formatado else '')
        
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor)
        
        linha_atual += 1
    
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    caminho_markdown = "sped_bloco_c.md"
    
    print("=" * 60)
    print("Parser Genérico do Bloco C - SPED EFD Contribuições")
    print("=" * 60)
    print(f"\nLendo definições do arquivo: {caminho_markdown}")
    
    definicoes = ler_definicoes_registros(caminho_markdown)
    
    if not definicoes:
        print("\nErro: Nenhuma definição de registro encontrada no arquivo markdown!")
        exit(1)
    
    print(f"\n{len(definicoes)} registros encontrados no markdown")
    
    # Lê os dados do bloco 0 uma vez para usar em todas as planilhas
    print("\nLendo dados do bloco 0...")
    dados_bloco_0 = ler_dados_bloco_0()
    if dados_bloco_0['cnpj']:
        print(f"CNPJ encontrado: {dados_bloco_0['cnpj']}")
    if dados_bloco_0['ie']:
        print(f"Inscrição Estadual encontrada: {dados_bloco_0['ie']}")
    if dados_bloco_0['periodo']:
        print(f"Período encontrado: {dados_bloco_0['periodo']}")
    
    print("\nProcessando arquivos...")
    print("=" * 60)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    for codigo_registro, campos_definicao in definicoes.items():
        caminho_arquivo = f"blocos_separados/bloco_C/{codigo_registro}.txt"
        
        print(f"\nProcessando registro {codigo_registro}...")
        registros = ler_arquivo_registro(caminho_arquivo, codigo_registro, campos_definicao)
        
        if registros:
            nome_arquivo_excel = f"registro_{codigo_registro}_{timestamp}.xlsx"
            criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo_excel, dados_bloco_0)
        else:
            print(f"Nenhum registro {codigo_registro} encontrado nos arquivos de entrada")
    
    print("\n" + "=" * 60)
    print("Processamento concluído!")
    print("=" * 60)

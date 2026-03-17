from pathlib import Path
import os
import re
import json
from datetime import datetime
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter


def ler_definicoes_registros(arquivo_markdown):
    """
    Lê o arquivo markdown e extrai as definições de todos os registros.
    Formato: markdown com tabelas no formato | Nº | Campo | Descrição | Tipo | Tam | Dec | Obrig |
    """
    definicoes = OrderedDict()
    
    if not os.path.exists(arquivo_markdown):
        print(f"Erro: Arquivo '{arquivo_markdown}' não encontrado!")
        return definicoes
    
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_markdown, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding do markdown detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo markdown!")
        return definicoes
    
    try:
        conteudo = arquivo.read()
    finally:
        arquivo.close()
    
    # Busca por padrões de registro: "REGISTRO 0000:" ou "## REGISTRO 0000:"
    padrao_registro = r'(?:##\s+)?REGISTRO\s+(\d{4}):'
    
    matches = list(re.finditer(padrao_registro, conteudo, re.IGNORECASE))
    
    for i, match in enumerate(matches):
        codigo_registro = match.group(1)
        inicio = match.end()
        
        if i + 1 < len(matches):
            fim = matches[i + 1].start()
        else:
            fim = len(conteudo)
        
        secao = conteudo[inicio:fim]
        
        padrao_secao_campos = r'###\s+Campos\s+do\s+Registro'
        match_secao = re.search(padrao_secao_campos, secao, re.IGNORECASE)
        
        if not match_secao:
            padrao_secao_campos = r'###\s+Campos'
            match_secao = re.search(padrao_secao_campos, secao, re.IGNORECASE)
        
        formato_texto_simples = False
        if not match_secao:
            # Busca pelo cabeçalho da tabela em formato texto simples
            padrao_cabecalho_texto = r'N[°ºoO]\s+Campo\s+Descri[çc][ãa]o\s+Tipo\s+Tam\s+Dec\s*Obrig'
            match_cabecalho_texto = re.search(padrao_cabecalho_texto, secao, re.IGNORECASE)
            if match_cabecalho_texto:
                match_secao = match_cabecalho_texto
                formato_texto_simples = True
        
        # Se ainda não encontrou, tenta buscar diretamente por linhas que começam com números de campo
        if not match_secao:
            # Procura por padrão de linha de campo: "01 REG ..."
            padrao_linha_campo = r'^\d{2}\s+\w+\s+'
            linhas_secao = secao.split('\n')
            for idx, linha in enumerate(linhas_secao):
                if re.match(padrao_linha_campo, linha.strip()):
                    # Encontrou uma linha de campo, assume que é formato texto simples
                    match_secao = re.search(r'.', secao)  # Match qualquer coisa para ter um objeto match
                    formato_texto_simples = True
                    break
        
        if match_secao or formato_texto_simples:
            inicio_tabela = match_secao.end()
            linhas_tabela = secao[inicio_tabela:].split('\n')
            
            campos = []
            em_tabela = False
            linha_cabecalho_encontrada = False
            
            if formato_texto_simples:
                # Se não encontrou o cabeçalho, usa toda a seção
                if not match_secao:
                    linhas_tabela = secao.split('\n')
                campos = _parsear_tabela_texto_simples(linhas_tabela)
                if campos:
                    definicoes[codigo_registro] = campos
                    print(f"Registro {codigo_registro} encontrado com {len(campos)} campos (formato texto simples)")
                continue
            
            for linha in linhas_tabela:
                linha = linha.strip()
                
                if (linha.startswith('###') or linha.startswith('##') or 
                    linha.startswith('---') or linha.startswith('REGISTRO')):
                    if campos:
                        break
                    continue
                
                if '|' in linha and ('Nº' in linha or 'Campo' in linha):
                    if re.search(r'\|\s*N[°ºoO]?\s*\|', linha, re.IGNORECASE) and 'Campo' in linha:
                        linha_cabecalho_encontrada = True
                        em_tabela = True
                        continue
                
                if em_tabela and linha.startswith('|') and re.match(r'^\|[\s\-:]+\|', linha):
                    continue
                
                if em_tabela and linha.startswith('|') and linha_cabecalho_encontrada:
                    partes = [p.strip() for p in linha.split('|') if p.strip()]
                    
                    if len(partes) >= 7:
                        try:
                            numero_campo = partes[0].strip()
                            nome_campo = partes[1].strip()
                            descricao = partes[2].strip()
                            tipo = partes[3].strip()
                            tamanho = partes[4].strip()
                            decimais = partes[5].strip()
                            obrigatorio = partes[6].strip()
                            
                            if not re.match(r'^\d{2}$', numero_campo):
                                continue
                            
                            descricao = re.sub(r'<br\s*/?>', ' ', descricao, flags=re.IGNORECASE)
                            descricao = re.sub(r'<[^>]+>', '', descricao)
                            descricao = descricao.strip()
                            
                            campo = {
                                'numero': numero_campo,
                                'campo': nome_campo,
                                'descricao': descricao,
                                'tipo': tipo,
                                'tamanho': tamanho,
                                'decimais': decimais,
                                'obrigatorio': obrigatorio
                            }
                            campos.append(campo)
                        except (IndexError, ValueError) as e:
                            continue
                    elif len(partes) >= 2:
                        try:
                            numero_campo = partes[0].strip()
                            
                            if not re.match(r'^\d{2}$', numero_campo):
                                continue
                            
                            nome_campo = partes[1].strip() if len(partes) > 1 else ''
                            
                            tipo = ''
                            tamanho = ''
                            decimais = ''
                            obrigatorio = ''
                            descricao_parts = []
                            
                            for i in range(len(partes) - 1, 1, -1):
                                parte = partes[i].strip()
                                if parte in ['C', 'N'] and i + 1 < len(partes):
                                    tipo = parte
                                    if i + 1 < len(partes):
                                        tamanho = partes[i + 1].strip()
                                    if i + 2 < len(partes):
                                        decimais = partes[i + 2].strip()
                                    if i + 3 < len(partes):
                                        obrigatorio = partes[i + 3].strip()
   
                                    descricao_parts = partes[2:i]
                                    break
                            
                            if not tipo and len(partes) >= 4:
                                ultimas = partes[-4:]
                                if len(ultimas) == 4:
                                    tipo_candidato = ultimas[0].strip()
                                    if tipo_candidato in ['C', 'N']:
                                        tipo = tipo_candidato
                                        tamanho = ultimas[1].strip()
                                        decimais = ultimas[2].strip()
                                        obrigatorio = ultimas[3].strip()
                                        descricao_parts = partes[2:-4]
                            
                            if not tipo:
                                texto_completo = ' '.join(partes[2:])
                                padrao_metadados = r'([CN])\s+(\d+\*?|-)\s+(-|\d+)\s+([A-Z]+)'
                                match_meta = re.search(padrao_metadados, texto_completo)
                                if match_meta:
                                    tipo = match_meta.group(1)
                                    tamanho = match_meta.group(2)
                                    decimais = match_meta.group(3)
                                    obrigatorio = match_meta.group(4)
                                    descricao_texto = re.sub(padrao_metadados, '', texto_completo).strip()
                                    descricao_parts = [descricao_texto] if descricao_texto else []
                            
                            descricao = ' '.join(descricao_parts).strip()
                            descricao = re.sub(r'<br\s*/?>', ' ', descricao, flags=re.IGNORECASE)
                            descricao = re.sub(r'<[^>]+>', '', descricao)
                            descricao = descricao.strip()
                            
                            if tipo in ['C', 'N']:
                                campo = {
                                    'numero': numero_campo,
                                    'campo': nome_campo,
                                    'descricao': descricao,
                                    'tipo': tipo,
                                    'tamanho': tamanho,
                                    'decimais': decimais,
                                    'obrigatorio': obrigatorio
                                }
                                campos.append(campo)
                        except Exception as e:
                            continue
                elif em_tabela and not linha.startswith('|') and linha:
                    if campos:
                        break
                    em_tabela = False
            
            if campos:
                definicoes[codigo_registro] = campos
                print(f"Registro {codigo_registro} encontrado com {len(campos)} campos")
    
    print(f"\nTotal de registros encontrados: {len(definicoes)}")
    return definicoes


def _parsear_tabela_texto_simples(linhas):
    """
    Parseia tabela em formato texto simples (sem markdown).
    Formato: "01 REG Texto fixo contendo "0500" C 004* - O"
    ou formato com descrição em múltiplas linhas:
    "02 COD_VER Código da versão do leiaute conforme a tabela indicada no N 003* - O"
    """
    campos = []
    
    for i, linha in enumerate(linhas):
        linha = linha.strip()
        
        # Para quando encontra seções que indicam fim da tabela
        if (linha.startswith('Observações') or linha.startswith('Observacoes') or 
            linha.startswith('Nível') or linha.startswith('Nivel') or 
            linha.startswith('###') or linha.startswith('##') or
            linha.startswith('---') or (linha.startswith('REGISTRO') and i > 0)):
            if campos:
                break
            continue
        
        # Ignora linhas vazias ou que são apenas números de página
        if not linha or re.match(r'^Página \d+', linha):
            continue
        
        # Padrão completo em uma linha: "01 REG Texto C 004 - O"
        padrao_completo = r'^(\d{2})\s+(\w+)\s+(.+?)\s+([CN])\s+(\d+\*?|-)\s+(-|\d+)\s+([A-Z]+)$'
        match_completo = re.match(padrao_completo, linha)
        
        if match_completo:
            numero_campo = match_completo.group(1)
            nome_campo = match_completo.group(2)
            descricao = match_completo.group(3).strip()
            tipo = match_completo.group(4)
            tamanho = match_completo.group(5)
            decimais = match_completo.group(6)
            obrigatorio = match_completo.group(7)
            
            # Limpa a descrição
            descricao = re.sub(r'["""]', '"', descricao)
            descricao = descricao.strip()
            
            campo = {
                'numero': numero_campo,
                'campo': nome_campo,
                'descricao': descricao,
                'tipo': tipo,
                'tamanho': tamanho,
                'decimais': decimais,
                'obrigatorio': obrigatorio
            }
            campos.append(campo)
            continue
        
        # Padrão com início: "01 REG ..." e precisa encontrar o final
        padrao_inicio = r'^(\d{2})\s+(\w+)\s+(.+)$'
        match_inicio = re.match(padrao_inicio, linha)
        
        if match_inicio:
            numero_campo = match_inicio.group(1)
            nome_campo = match_inicio.group(2)
            resto = match_inicio.group(3).strip()
            
            # Tenta encontrar padrão no final: "C 004 - O" ou "N 003* - O"
            padrao_fim = r'^(.+?)\s+([CN])\s+(\d+\*?|-)\s+(-|\d+)\s+([A-Z]+)$'
            match_fim = re.match(padrao_fim, resto)
            
            if match_fim:
                descricao = match_fim.group(1).strip()
                tipo = match_fim.group(2)
                tamanho = match_fim.group(3)
                decimais = match_fim.group(4)
                obrigatorio = match_fim.group(5)
                
                # Limpa a descrição
                descricao = re.sub(r'["""]', '"', descricao)
                descricao = descricao.strip()
                
                campo = {
                    'numero': numero_campo,
                    'campo': nome_campo,
                    'descricao': descricao,
                    'tipo': tipo,
                    'tamanho': tamanho,
                    'decimais': decimais,
                    'obrigatorio': obrigatorio
                }
                campos.append(campo)
                continue
            
            # Se não encontrou o padrão completo, tenta buscar nas próximas linhas
            # ou usar regex mais flexível
            padrao_flexivel = r'^(.+?)\s+([CN])\s+(\d+\*?|-)\s+(-|\d+)\s+([A-Z]+(?:\s+[A-Z]+)?)$'
            match_flexivel = re.search(padrao_flexivel, resto)
            
            if match_flexivel:
                descricao = resto[:match_flexivel.start()].strip()
                tipo = match_flexivel.group(2)
                tamanho = match_flexivel.group(3)
                decimais = match_flexivel.group(4)
                obrigatorio = match_flexivel.group(5).strip()
                
                # Limpa a descrição
                descricao = re.sub(r'["""]', '"', descricao)
                descricao = descricao.strip()
                
                if descricao and tipo in ['C', 'N']:
                    campo = {
                        'numero': numero_campo,
                        'campo': nome_campo,
                        'descricao': descricao,
                        'tipo': tipo,
                        'tamanho': tamanho,
                        'decimais': decimais,
                        'obrigatorio': obrigatorio
                    }
                    campos.append(campo)
    
    return campos


def parse_registro_generico(linha, codigo_registro, campos_definicao):
    """
    Parseia uma linha do SPED de forma genérica baseado nas definições.
    """
    linha = linha.strip()
    if not linha:
        return None
    
    if f"|{codigo_registro}|" not in linha and not linha.startswith(f"|{codigo_registro}|"):
        return None
    
    partes = linha.split('|')
    
    if partes and not partes[0]:
        partes = partes[1:]
    
    if partes and not partes[-1]:
        partes = partes[:-1]
    
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip() == codigo_registro:
                idx = i
                break
        
        if idx is None:
            return None
        
        campos = {}
        for campo_def in campos_definicao:
            numero_campo = int(campo_def['numero'])
            nome_campo = campo_def['campo']
            
            posicao = idx + (numero_campo - 1)
            
            if posicao < len(partes):
                valor = partes[posicao].strip()
                campos[nome_campo] = valor
            else:
                campos[nome_campo] = ''
        
        return campos
        
    except (IndexError, ValueError, KeyError) as e:
        print(f"Erro ao parsear linha do registro {codigo_registro}: {e}")
        return None


def formatar_cnpj(cnpj_str):
    """Formata CNPJ para o formato XX.XXX.XXX/XXXX-XX."""
    if not cnpj_str or len(cnpj_str) != 14:
        return cnpj_str
    
    try:
        return f"{cnpj_str[0:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:14]}"
    except:
        return cnpj_str


def formatar_cpf(cpf_str):
    """Formata CPF para o formato XXX.XXX.XXX-XX."""
    if not cpf_str or len(cpf_str) != 11:
        return cpf_str
    
    try:
        return f"{cpf_str[0:3]}.{cpf_str[3:6]}.{cpf_str[6:9]}-{cpf_str[9:11]}"
    except:
        return cpf_str


def formatar_cep(cep_str):
    """Formata CEP para o formato XXXXX-XXX."""
    if not cep_str or len(cep_str) != 8:
        return cep_str
    
    try:
        return f"{cep_str[0:5]}-{cep_str[5:8]}"
    except:
        return cep_str


def formatar_telefone(tel_str):
    """Formata telefone para o formato (XX) XXXXX-XXXX ou (XX) XXXX-XXXX."""
    if not tel_str:
        return tel_str
    
    try:
        tel_limpo = ''.join(filter(str.isdigit, tel_str))
        if len(tel_limpo) == 10:
            return f"({tel_limpo[0:2]}) {tel_limpo[2:6]}-{tel_limpo[6:10]}"
        elif len(tel_limpo) == 11:
            return f"({tel_limpo[0:2]}) {tel_limpo[2:7]}-{tel_limpo[7:11]}"
        else:
            return tel_str
    except:
        return tel_str


def formatar_data(data_str):
    """Formata data do formato DDMMYYYY para DD/MM/YYYY."""
    if not data_str or len(data_str) != 8:
        return data_str
    
    try:
        return f"{data_str[0:2]}/{data_str[2:4]}/{data_str[4:8]}"
    except:
        return data_str


def aplicar_formatacao(campo, valor, tipo_campo):
    """Aplica formatação específica baseada no tipo e nome do campo."""
    if not valor:
        return valor
    
    if campo == 'CNPJ' and len(valor) == 14:
        return formatar_cnpj(valor)
    
    if campo == 'CPF' and len(valor) == 11:
        return formatar_cpf(valor)
    
    if campo == 'CEP' and len(valor) == 8:
        return formatar_cep(valor)
    
    if campo in ['FONE', 'FAX']:
        return formatar_telefone(valor)
    
    if campo.startswith('DT_') or campo.startswith('DTE_'):
        if len(valor) == 8 and valor.isdigit():
            return formatar_data(valor)
    
    return valor


def ler_arquivo_registro(caminho_arquivo, codigo_registro, campos_definicao):
    """
    Lê o arquivo de um registro específico e retorna os registros parseados.
    """
    registros = []
    
    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo '{caminho_arquivo}' não encontrado")
        return registros
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo '{caminho_arquivo}'!")
        return registros
    
    try:
        for linha in arquivo:
            registro = parse_registro_generico(linha, codigo_registro, campos_definicao)
            if registro:
                registros.append(registro)
    finally:
        if arquivo:
            arquivo.close()
    
    return registros


def criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo=None):
    """
    Cria um arquivo Excel (.xlsx) com os registros parseados.
    """
    if not registros:
        print(f"Nenhum registro {codigo_registro} encontrado!")
        return None
    
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"registro_{codigo_registro}_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {codigo_registro}"
    
    cabecalhos = []
    for campo in campos_definicao:
        numero = campo['numero'].zfill(2) if campo['numero'].isdigit() else campo['numero']
        cabecalho = f"{numero} - {campo['campo']}"
        cabecalhos.append(cabecalho)
    
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    linha_atual = 2
    
    for registro in registros:
        valores = []
        for campo_def in campos_definicao:
            nome_campo = campo_def['campo']
            tipo_campo = campo_def['tipo']
            valor = registro.get(nome_campo, '')
            
            valor_formatado = aplicar_formatacao(nome_campo, valor, tipo_campo)
            valores.append(valor_formatado if valor_formatado else '')
        
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor)
        
        linha_atual += 1
    
    for col_idx in range(1, len(cabecalhos) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    try:
        wb.save(nome_arquivo)
        print(f"Arquivo Excel criado: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def ler_bloco_0_do_sped(caminho_arquivo_sped):
    """
    Lê o arquivo SPED e extrai todas as linhas do bloco 0.
    Retorna uma lista de linhas do bloco 0.
    """
    linhas_bloco_0 = []
    
    if not os.path.exists(caminho_arquivo_sped):
        print(f"Erro: Arquivo SPED '{caminho_arquivo_sped}' não encontrado!")
        return linhas_bloco_0
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo_sped, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo SPED!")
        return linhas_bloco_0
    
    try:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            
            # Verifica se é uma linha do bloco 0 (registros que começam com 0)
            # Formato: |0000|... ou |0001|... etc
            if linha.startswith('|') and '|' in linha[1:]:
                partes = linha.split('|')
                if len(partes) > 1 and partes[1] and partes[1].startswith('0'):
                    linhas_bloco_0.append(linha)
    finally:
        if arquivo:
            arquivo.close()
    
    return linhas_bloco_0


def extrair_datas_do_registro_0000(linhas_bloco_0):
    """
    Extrai as datas DT_INI e DT_FIN do registro 0000.
    Retorna (data_ini_str, data_fin_str) no formato DDMMYYYY ou (None, None) se não encontrar.
    """
    for linha in linhas_bloco_0:
        if '|0000|' in linha or linha.startswith('|0000|'):
            partes = linha.split('|')
            
            # O formato do SPED tem pipes no início e fim, então:
            # partes[0] = '' (antes do primeiro |)
            # partes[1] = '0000' (REG)
            # partes[2] = COD_VER
            # partes[3] = COD_FIN
            # partes[4] = DT_INI
            # partes[5] = DT_FIN
            # partes[-1] = '' (após o último |)
            
            if len(partes) >= 6:
                # Verifica se realmente é o registro 0000
                if partes[1].strip() == '0000':
                    dt_ini_str = partes[4].strip() if len(partes) > 4 else ''
                    dt_fin_str = partes[5].strip() if len(partes) > 5 else ''
                    
                    if dt_ini_str and dt_fin_str and len(dt_ini_str) == 8 and len(dt_fin_str) == 8:
                        return dt_ini_str, dt_fin_str
    
    return None, None


def extrair_mes_ano_da_data(data_ddmmyyyy):
    """
    Extrai mês e ano de uma data no formato DDMMYYYY.
    Retorna (ano, mes) como inteiros.
    """
    if not data_ddmmyyyy or len(data_ddmmyyyy) != 8:
        return None, None
    
    try:
        ano = int(data_ddmmyyyy[4:8])
        mes = int(data_ddmmyyyy[2:4])
        return ano, mes
    except:
        return None, None


def extrair_mes_ano_da_data_criacao(data_criacao):
    """
    Extrai mês e ano de uma data no formato YYYY-MM-DD.
    Retorna (ano, mes) como inteiros.
    """
    if not data_criacao:
        return None, None
    
    try:
        partes = data_criacao.split('-')
        if len(partes) >= 2:
            ano = int(partes[0])
            mes = int(partes[1])
            return ano, mes
    except:
        pass
    
    return None, None


def buscar_versao_por_mes_ano(ano_arquivo, mes_arquivo, caminho_versoes_json):
    """
    Busca a versão apropriada no arquivo versoes.json baseado no mês/ano.
    
    Lógica:
    1. Busca versão com mês/ano igual ao do arquivo
    2. Se não encontrar, busca o mês mais avançado (anterior) ao período do arquivo
       (exemplo: arquivo 02/2026, só há 01/2026 e 03/2026, usa 01/2026)
    
    Retorna a versão encontrada ou None.
    """
    if not os.path.exists(caminho_versoes_json):
        print(f"Erro: Arquivo '{caminho_versoes_json}' não encontrado!")
        return None
    
    with open(caminho_versoes_json, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    versoes = dados.get("versoes", [])
    
    if not versoes:
        print("Erro: Nenhuma versão encontrada no arquivo versoes.json!")
        return None
    
    # Primeiro, tenta encontrar versão com mês/ano exato
    versao_exata = None
    versoes_anteriores = []
    
    for versao in versoes:
        data_criacao = versao.get("data_criacao", "")
        ano_versao, mes_versao = extrair_mes_ano_da_data_criacao(data_criacao)
        
        if ano_versao is None or mes_versao is None:
            continue
        
        # Verifica se é o mês/ano exato
        if ano_versao == ano_arquivo and mes_versao == mes_arquivo:
            versao_exata = versao
            break
        
        # Coleta versões anteriores (mesmo ano com mês menor, ou ano anterior)
        if (ano_versao < ano_arquivo) or (ano_versao == ano_arquivo and mes_versao < mes_arquivo):
            versoes_anteriores.append((versao, ano_versao, mes_versao))
    
    if versao_exata:
        print(f"Versão encontrada (mês/ano exato): {versao_exata['versao']} (data_criacao: {versao_exata['data_criacao']})")
        return versao_exata
    
    # Se não encontrou exato, busca o mês mais avançado (anterior) mais próximo
    if versoes_anteriores:
        # Ordena por ano e mês (mais recente primeiro)
        versoes_anteriores.sort(key=lambda x: (x[1], x[2]), reverse=True)
        versao_encontrada = versoes_anteriores[0][0]
        print(f"Versão encontrada (mês anterior mais próximo): {versao_encontrada['versao']} (data_criacao: {versao_encontrada['data_criacao']})")
        print(f"  Período do arquivo: {mes_arquivo:02d}/{ano_arquivo}")
        return versao_encontrada
    
    # Se não encontrou nenhuma versão anterior, usa a mais recente disponível
    versoes_ordenadas = sorted(versoes, key=lambda x: x.get("data_criacao", ""), reverse=True)
    if versoes_ordenadas:
        versao_encontrada = versoes_ordenadas[0]
        print(f"Aviso: Nenhuma versão anterior encontrada. Usando a mais recente: {versao_encontrada['versao']} (data_criacao: {versao_encontrada['data_criacao']})")
        return versao_encontrada
    
    return None


def processar_bloco_0(sped_arquivo=None):
    """
    Função principal que processa o bloco 0 do arquivo SPED.
    
    Fluxo:
    1. Lê o arquivo SPED e extrai linhas do bloco 0
    2. Identifica DT_INI e DT_FIN no registro 0000
    3. Busca a versão apropriada no versoes.json (por mês/ano)
    4. Carrega a documentação markdown da versão encontrada
    5. Processa todos os registros do bloco 0
    6. Gera arquivos Excel para cada tipo de registro
    """
    base_dir = Path(__file__).parent
    
    # Se não foi fornecido um arquivo, usa o padrão
    if sped_arquivo is None:
        sped_arquivo = "08589276000169-9084073269-20201101-20201130-0-911F0BE8FE2F193E99CEFD2FA579B4E7F40C8ED9-SPED-EFD.txt"
    
    caminho_sped = base_dir / sped_arquivo if not os.path.isabs(sped_arquivo) else Path(sped_arquivo)
    caminho_versoes_json = base_dir / "versoes.json"
    
    print("=" * 60)
    print("Parser do Bloco 0 - SPED EFD Fiscal")
    print("=" * 60)
    print(f"\nLendo arquivo SPED: {caminho_sped}")
    
    # 1. Ler o bloco 0 do arquivo SPED
    linhas_bloco_0 = ler_bloco_0_do_sped(str(caminho_sped))
    
    if not linhas_bloco_0:
        print("Erro: Nenhuma linha do bloco 0 encontrada no arquivo SPED!")
        return
    
    print(f"Encontradas {len(linhas_bloco_0)} linhas do bloco 0")
    
    # 2. Extrair datas DT_INI e DT_FIN do registro 0000
    dt_ini_str, dt_fin_str = extrair_datas_do_registro_0000(linhas_bloco_0)
    
    if not dt_ini_str or not dt_fin_str:
        print("Erro: Não foi possível extrair as datas DT_INI e DT_FIN do registro 0000!")
        return
    
    print(f"\nDatas extraídas do registro 0000:")
    print(f"  DT_INI: {dt_ini_str} (DDMMYYYY)")
    print(f"  DT_FIN: {dt_fin_str} (DDMMYYYY)")
    
    # Extrai mês/ano da data final (usando DT_FIN como referência)
    ano_arquivo, mes_arquivo = extrair_mes_ano_da_data(dt_fin_str)
    
    if not ano_arquivo or not mes_arquivo:
        print("Erro: Não foi possível extrair mês/ano das datas!")
        return
    
    print(f"  Período do arquivo: {mes_arquivo:02d}/{ano_arquivo}")
    
    # 3. Buscar versão correspondente no versoes.json
    print(f"\nBuscando versão correspondente em {caminho_versoes_json}...")
    versao_info = buscar_versao_por_mes_ano(ano_arquivo, mes_arquivo, str(caminho_versoes_json))
    
    if not versao_info:
        print("Erro: Não foi possível encontrar uma versão válida!")
        return
    
    versao = versao_info["versao"]
    
    # 4. Buscar arquivo markdown da documentação do bloco 0
    caminho_markdown = base_dir / "documentacao_blocos" / versao / "bloco_0.md"
    
    if not caminho_markdown.exists():
        print(f"Erro: Arquivo de documentação não encontrado: {caminho_markdown}")
        return
    
    print(f"\nUsando documentação: {caminho_markdown}")
    print(f"Lendo definições dos registros...")
    
    # 5. Ler definições dos registros do markdown
    definicoes = ler_definicoes_registros(str(caminho_markdown))
    
    if not definicoes:
        print("\nErro: Nenhuma definição de registro encontrada no arquivo markdown!")
        return
    
    print(f"\n{len(definicoes)} registros encontrados no markdown")
    print("\nProcessando registros do bloco 0...")
    print("=" * 60)
    
    # 6. Processar cada tipo de registro encontrado no bloco 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    registros_processados = {}
    
    # Agrupa linhas por tipo de registro
    for linha in linhas_bloco_0:
        partes = linha.split('|')
        
        # O primeiro campo após o pipe inicial é o código do registro
        if len(partes) > 1:
            codigo_registro = partes[1].strip()
            
            if codigo_registro:
                if codigo_registro not in registros_processados:
                    registros_processados[codigo_registro] = []
                
                registros_processados[codigo_registro].append(linha)
    
    # Processa cada tipo de registro
    for codigo_registro, linhas_registro in registros_processados.items():
        if codigo_registro not in definicoes:
            print(f"\nAviso: Registro {codigo_registro} não encontrado nas definições do markdown. Pulando...")
            continue
        
        print(f"\nProcessando Registro {codigo_registro} ({len(linhas_registro)} ocorrências)")
        
        campos_definicao = definicoes[codigo_registro]
        
        # Parseia todas as linhas deste registro
        registros_parseados = []
        for linha in linhas_registro:
            registro = parse_registro_generico(linha, codigo_registro, campos_definicao)
            if registro:
                registros_parseados.append(registro)
        
        if registros_parseados:
            nome_arquivo_excel = base_dir / f"registro_{codigo_registro}_{timestamp}.xlsx"
            criar_arquivo_excel_registro(registros_parseados, codigo_registro, campos_definicao, str(nome_arquivo_excel))
        else:
            print(f"  Nenhum registro {codigo_registro} válido encontrado")
    
    print("\n" + "=" * 60)
    print("Processamento concluído!")
    print("=" * 60)


if __name__ == "__main__":
    sped_arquivo = "38045115000841-0390189812-20250201-20250228-0-769A2231EC07CDAA66A10B417EC8F709B8A1894D-SPED-EFD.txt"
    processar_bloco_0(sped_arquivo)


'tipo_operacao', 'indicador_emitente'

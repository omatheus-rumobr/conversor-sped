from pathlib import Path
import os
import json
from datetime import datetime
from modulos.bloco_c.rc001 import validar_c001
from modulos.bloco_c.rc100 import validar_c100
from modulos.bloco_c.rc101 import validar_c101
from modulos.bloco_c.rc110 import validar_c110
from modulos.bloco_c.rc170 import validar_c170
from modulos.bloco_c.rc190 import validar_c190
from modulos.bloco_0.r0001 import validar_0001
from modulos.bloco_0.r0005 import validar_0005
from modulos.bloco_0.r0150 import validar_0150_fiscal

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook

# Importa funções de formatação dos utils
try:
    from utils.cnpj import formatar_cnpj
    from utils.data_hora import formatar_data
    from utils.moeda import formatar_valor_monetario
except ImportError:
    # Fallback caso os utils não existam
    def formatar_cnpj(cnpj):
        if len(cnpj) == 14 and cnpj.isdigit():
            return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
        return cnpj
    
    def formatar_data(data):
        if len(data) == 8 and data.isdigit():
            return f"{data[:2]}/{data[2:4]}/{data[4:]}"
        return data
    
    def formatar_valor_monetario(valor):
        if not valor:
            return valor
        try:
            return str(valor).replace(',', '.')
        except:
            return valor


def validar_linha_sped(linha):
    """
    Valida uma linha do SPED substituindo "||" (dois pipes consecutivos) por "|-|".
    Faz substituição recursiva para garantir que todas as ocorrências sejam tratadas.
    
    Args:
        linha: String com a linha do SPED
        
    Returns:
        String com a linha validada (com "||" substituído por "|-|")
    """
    linha_validada = linha
    while '||' in linha_validada:
        linha_validada = linha_validada.replace('||', '|-|')
    return linha_validada


def ler_linhas_bloco_0(arquivo_sped):
    """
    Lê um arquivo SPED (.txt) e separa todas as linhas do bloco 0 que começam com "|0*|".
    
    Args:
        arquivo_sped: Caminho para o arquivo SPED (.txt)
        
    Returns:
        Lista de strings contendo as linhas do bloco 0
    """
    linhas_bloco_0 = []
    
    if not os.path.exists(arquivo_sped):
        print(f"Erro: Arquivo SPED '{arquivo_sped}' não encontrado!")
        return linhas_bloco_0
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_sped, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo SPED!")
        return linhas_bloco_0
    
    try:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            
            if linha.startswith('|0') and '|' in linha[2:]:
                partes = linha.split('|')
                if len(partes) >= 2 and partes[1].startswith('0'):
                    linha_validada = validar_linha_sped(linha)
                    linhas_bloco_0.append(linha_validada)
    finally:
        if arquivo:
            arquivo.close()
    
    return linhas_bloco_0


def ler_linhas_bloco_c(arquivo_sped):
    """
    Lê um arquivo SPED (.txt) e separa todas as linhas do bloco C que começam com "|C*|".
    
    Args:
        arquivo_sped: Caminho para o arquivo SPED (.txt)
        
    Returns:
        Lista de strings contendo as linhas do bloco C
    """
    linhas_bloco_c = []
    
    if not os.path.exists(arquivo_sped):
        print(f"Erro: Arquivo SPED '{arquivo_sped}' não encontrado!")
        return linhas_bloco_c
    

    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_sped, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo SPED!")
        return linhas_bloco_c
    
    try:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            
            if linha.startswith('|C') and '|' in linha[2:]:
                partes = linha.split('|')
                if len(partes) >= 2 and partes[1].startswith('C'):
                    linha_validada = validar_linha_sped(linha)
                    linhas_bloco_c.append(linha_validada)
    finally:
        if arquivo:
            arquivo.close()
    
    return linhas_bloco_c


def processar_json_registro(json_str):
    """
    Processa o JSON retornado pelos módulos de validação e retorna lista de dicionários simplificados.
    
    Args:
        json_str: String JSON retornada pela função de validação
        
    Returns:
        Lista de dicionários no formato {'CAMPO': 'valor', ...}
    """
    if not json_str:
        return []
    
    try:
        dados = json.loads(json_str)
        
        if isinstance(dados, dict):
            dados = [dados]
        
        registros_simplificados = []
        for registro in dados:
            registro_simplificado = {}
            for campo, info in registro.items():
                if isinstance(info, dict):
                    valor = info.get('valor', '')
                    registro_simplificado[campo] = valor
                else:
                    registro_simplificado[campo] = info
            registros_simplificados.append(registro_simplificado)
        
        return registros_simplificados
    except json.JSONDecodeError as e:
        print(f"Erro ao processar JSON: {e}")
        return []


def obter_cabecalhos_do_json(json_str):
    """
    Extrai os títulos dos campos do JSON para usar como cabeçalhos do Excel.
    
    Args:
        json_str: String JSON retornada pela função de validação
        
    Returns:
        Lista de tuplas (nome_campo, titulo_campo) ordenada
    """
    if not json_str:
        return []
    
    try:
        dados = json.loads(json_str)
        
        if isinstance(dados, dict):
            dados = [dados]
        
        if not dados:
            return []
        
        primeiro_registro = dados[0]
        cabecalhos = []
        
        for campo, info in primeiro_registro.items():
            if isinstance(info, dict):
                titulo = info.get('titulo', campo)
                cabecalhos.append((campo, titulo))
            else:
                cabecalhos.append((campo, campo))
        
        return cabecalhos
    except json.JSONDecodeError:
        return []


def obter_descricao_campo(campo, valor):
    """
    Retorna descrição formatada para campos específicos baseado no valor.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        
    Returns:
        String com descrição formatada ou valor original
    """
    if not valor:
        return valor
    
    descricoes = {
        'IND_MOV': {
            '0': '0 - Bloco com dados informados',
            '1': '1 - Bloco sem dados informados'
        },
        'IND_OPER': {
            '0': '0 - Entrada',
            '1': '1 - Saída'
        },
        'IND_EMIT': {
            '0': '0 - Emissão própria',
            '1': '1 - Terceiros'
        },
        'IND_PGTO': {
            '0': '0 - À vista',
            '1': '1 - A prazo',
            '2': '2 - Sem indicação',
            '9': '9 - Sem pagamento'
        },
        'IND_FRT': {
            '0': '0 - Por conta do remetente (CIF)',
            '1': '1 - Por conta do destinatário (FOB)',
            '2': '2 - Por conta de terceiros',
            '9': '9 - Sem frete'
        }
    }
    
    if campo in descricoes and valor in descricoes[campo]:
        return descricoes[campo][valor]
    
    return valor


def aplicar_formatacao(campo, valor, tipo_campo):
    """
    Aplica formatação específica baseada no tipo e nome do campo.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        tipo_campo: Tipo do campo (C, N, etc)
        
    Returns:
        Valor formatado
    """
    if not valor:
        return valor
    
    if campo == 'CNPJ' and len(str(valor)) == 14:
        return formatar_cnpj(str(valor))
    
    if campo.startswith('DT_') or campo.startswith('DTE_'):
        valor_str = str(valor)
        if len(valor_str) == 8 and valor_str.isdigit():
            return formatar_data(valor_str)
    
    if campo.startswith('VL_') and tipo_campo == 'N':
        return formatar_valor_monetario(valor)
    
    return obter_descricao_campo(campo, valor)


def extrair_dados_bloco_0(linhas_bloco_0):
    """
    Extrai os dados do bloco 0 (CNPJ, IE, Período) usando as funções de validação.
    
    Args:
        linhas_bloco_0: Lista de linhas do bloco 0
        
    Returns:
        Dicionário com cnpj, ie e periodo
    """
    dados = {
        'cnpj': '',
        'ie': '',
        'periodo': ''
    }
    
    for linha in linhas_bloco_0:
        if '|0000|' in linha or linha.startswith('|0000|'):
            partes = linha.split('|')
            if len(partes) >= 8:
                if partes[1].strip() == '0000':
                    cnpj = partes[7].strip() if len(partes) > 7 else ''
                    if cnpj and cnpj != '-':
                        if len(cnpj) == 14 and cnpj.isdigit():
                            dados['cnpj'] = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                        else:
                            dados['cnpj'] = cnpj
                    
                    dt_ini = partes[4].strip() if len(partes) > 4 else ''
                    dt_fin = partes[5].strip() if len(partes) > 5 else ''
                    if dt_ini and dt_ini != '-' and dt_fin and dt_fin != '-' and len(dt_ini) == 8 and len(dt_fin) == 8:
                        periodo_ini = f"{dt_ini[:2]}/{dt_ini[2:4]}/{dt_ini[4:]}"
                        periodo_fin = f"{dt_fin[:2]}/{dt_fin[2:4]}/{dt_fin[4:]}"
                        dados['periodo'] = f"{periodo_ini} a {periodo_fin}"
                    
                    ie = partes[10].strip() if len(partes) > 10 else ''
                    if ie and ie != '-':
                        dados['ie'] = ie
                    break
    
    return dados


def buscar_nome_municipio(cod_mun):
    """
    Busca o nome do município no arquivo municipios.json usando o código.
    
    Args:
        cod_mun: Código do município (7 dígitos)
        
    Returns:
        String com o nome do município ou "Município não identificado" se não encontrado
    """
    if not cod_mun:
        return "Município não identificado"
    
    cod_mun_str = str(cod_mun).strip()
    
    if not cod_mun_str or cod_mun_str == '-':
        return "Município não identificado"
    
    try:
        base_dir = Path(__file__).parent
        caminho_municipios = base_dir / "municipios.json"
        
        if not caminho_municipios.exists():
            return "Município não identificado"
        
        with open(caminho_municipios, 'r', encoding='utf-8') as f:
            municipios = json.load(f)
        
        for municipio in municipios:
            codigo_municipio = str(municipio.get('codigo', '')).strip()
            if cod_mun_str == codigo_municipio:
                return municipio.get('nome', 'Município não identificado')
        
        cod_mun_limpo = cod_mun_str.lstrip('0')
        if cod_mun_limpo and cod_mun_limpo != cod_mun_str:
            for municipio in municipios:
                codigo_municipio = str(municipio.get('codigo', '')).strip()
                codigo_limpo = codigo_municipio.lstrip('0')
                if cod_mun_limpo == codigo_limpo:
                    return municipio.get('nome', 'Município não identificado')
        
        return "Município não identificado"
    except Exception as e:
        print(f"Erro ao buscar município: {e}")
        return "Município não identificado"


def extrair_dados_participantes(linhas_bloco_0):
    """
    Extrai os dados dos participantes do registro 0150 usando a função de validação.
    
    Args:
        linhas_bloco_0: Lista de linhas do bloco 0
        
    Returns:
        Dicionário com COD_PART como chave e dicionário com dados do participante como valor
    """
    participantes = {}
    
    linhas_0150 = [linha for linha in linhas_bloco_0 if linha.startswith('|0150|')]
    
    if not linhas_0150:
        return participantes
    
    json_resultado = validar_0150_fiscal(linhas_0150)
    
    if not json_resultado:
        return participantes
    
    try:
        dados = json.loads(json_resultado)
        
        if isinstance(dados, dict):
            dados = [dados]
        
        for registro in dados:
            cod_part = ''
            nome = ''
            cnpj = ''
            cpf = ''
            cod_mun = ''
            
            for campo, info in registro.items():
                if isinstance(info, dict):
                    valor = info.get('valor', '')
                    if campo == 'COD_PART':
                        cod_part = valor
                    elif campo == 'NOME':
                        nome = valor
                    elif campo == 'CNPJ':
                        cnpj = valor
                    elif campo == 'CPF':
                        cpf = valor
                    elif campo == 'COD_MUN':
                        cod_mun = valor
            
            if cod_part:
                cnpj_formatado = cnpj
                if cnpj and len(cnpj) == 14 and cnpj.isdigit():
                    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                
                cpf_formatado = cpf
                if cpf and len(cpf) == 11 and cpf.isdigit():
                    cpf_formatado = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                
                nome_municipio = buscar_nome_municipio(cod_mun)
                
                participantes[cod_part] = {
                    'cod_part': cod_part,
                    'cnpj': cnpj_formatado,
                    'cpf': cpf_formatado,
                    'nome': nome,
                    'cod_mun': cod_mun,
                    'nome_municipio': nome_municipio
                }
    except json.JSONDecodeError as e:
        print(f"Erro ao processar JSON dos participantes: {e}")
    
    return participantes


def extrair_cod_part_por_ordem(linhas_bloco_c, codigo_registro_filho):
    """
    Extrai o COD_PART do C100 pai para cada registro filho baseado na ordem sequencial.
    
    Args:
        linhas_bloco_c: Lista de todas as linhas do bloco C em ordem
        codigo_registro_filho: Código do registro filho (ex: 'C170', 'C190')
        
    Returns:
        Lista de COD_PART na mesma ordem dos registros filhos encontrados
    """
    cod_parts = []
    cod_part_atual = ''
    
    for linha in linhas_bloco_c:
        partes = linha.split('|')
        if len(partes) < 2:
            continue
        
        reg = partes[1].strip()
        
        if reg == 'C100' and len(partes) > 4:
            cod_part_atual = partes[4].strip() if partes[4].strip() else ''
        
        elif reg == codigo_registro_filho:
            cod_parts.append(cod_part_atual)
    
    return cod_parts


def criar_arquivo_excel_registro_json(registros_json, codigo_registro, nome_arquivo=None, dados_bloco_0=None, participantes=None, linhas_originais=None):
    """
    Cria um arquivo Excel (.xlsx) com os registros parseados a partir do JSON.
    
    Args:
        registros_json: String JSON retornada pela função de validação
        codigo_registro: Código do registro (ex: 'C001')
        nome_arquivo: Nome do arquivo Excel a ser criado (opcional)
        dados_bloco_0: Dicionário com cnpj, ie e periodo do bloco 0 (opcional)
        participantes: Dicionário com dados dos participantes (opcional)
        linhas_originais: Lista de linhas originais do bloco C para fazer o mapeamento com C100 (opcional)
    """
    registros = processar_json_registro(registros_json)
    
    if not registros:
        print(f"Nenhum registro {codigo_registro} encontrado!")
        return None
    
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"registro_{codigo_registro}_{timestamp}.xlsx"
    
    if dados_bloco_0 is None:
        dados_bloco_0 = {'cnpj': '', 'ie': '', 'periodo': ''}
    
    if participantes is None:
        participantes = {}
    
    cabecalhos_campos = obter_cabecalhos_do_json(registros_json)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {codigo_registro}"
    
    cabecalhos = ['CNPJ', 'Inscrição Estadual', 'Período']
    cabecalhos.extend(
        ['Código Participante', 'CNPJ Participante', 'CPF Participante', 'Nome Participante', 'Código Município', 'Nome Município']
    )
    cabecalhos.extend([titulo for _, titulo in cabecalhos_campos])
    
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    linha_atual = 2
    
    registros_filhos = ['C170', 'C190', 'C101', 'C105', 'C110', 'C111', 'C112', 'C113']
    usar_mapeamento = codigo_registro in registros_filhos
    
    cod_parts_pais = []
    if usar_mapeamento and linhas_originais:
        cod_parts_pais = extrair_cod_part_por_ordem(linhas_originais, codigo_registro)
    
    for idx_registro, registro in enumerate(registros):
        valores = []
        valores.append(dados_bloco_0.get('cnpj', ''))
        valores.append(dados_bloco_0.get('ie', ''))
        valores.append(dados_bloco_0.get('periodo', ''))
        
        cod_part = registro.get('COD_PART', '')
        
        if not cod_part and usar_mapeamento and cod_parts_pais and idx_registro < len(cod_parts_pais):
            cod_part = cod_parts_pais[idx_registro]
        
        if cod_part:
            if cod_part in participantes:
                participante = participantes[cod_part]
                valores.append(participante.get('cod_part', ''))
                valores.append(participante.get('cnpj', ''))
                valores.append(participante.get('cpf', ''))
                valores.append(participante.get('nome', ''))
                valores.append(participante.get('cod_mun', ''))
                valores.append(participante.get('nome_municipio', ''))
            else:
                valores.extend(['', '', '', '', '', ''])
        else:
            valores.extend(['', '', '', '', '', ''])
        
        for campo, _ in cabecalhos_campos:
            valor = registro.get(campo, '')
            
            tipo_campo = 'C'
            if campo.startswith('VL_') or campo.startswith('QTD') or campo.startswith('ALIQ'):
                tipo_campo = 'N'
            
            valor_formatado = aplicar_formatacao(campo, valor, tipo_campo)
            valores.append(valor_formatado if valor_formatado else '')
        
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor)
        
        linha_atual += 1
    
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


if __name__ == "__main__":
    arquivo_sped = "38045115000841-0390189812-20250201-20250228-0-769A2231EC07CDAA66A10B417EC8F709B8A1894D-SPED-EFD.txt"
    
    base_dir = Path(__file__).parent
    caminho_arquivo = base_dir / arquivo_sped if not os.path.isabs(arquivo_sped) else Path(arquivo_sped)
    
    print("=" * 60)
    print("Processador do Bloco C - SPED EFD Fiscal")
    print("=" * 60)
    print(f"\nLendo arquivo SPED: {caminho_arquivo}")
    
    linhas = ler_linhas_bloco_c(str(caminho_arquivo))
    print(f"Total de linhas do bloco C encontradas: {len(linhas)}")
    
    linhas_bloco_0 = ler_linhas_bloco_0(str(caminho_arquivo))
    dados_bloco_0 = extrair_dados_bloco_0(linhas_bloco_0)
    participantes = extrair_dados_participantes(linhas_bloco_0)
    
    print(f"\nCNPJ: {dados_bloco_0.get('cnpj', 'N/A')}")
    print(f"Inscrição Estadual: {dados_bloco_0.get('ie', 'N/A')}")
    print(f"Período: {dados_bloco_0.get('periodo', 'N/A')}")
    print(f"Participantes encontrados: {len(participantes)}")
    
    validadores = {
        'C001': (validar_c001, True),
        'C100': (validar_c100, False),
        'C101': (validar_c101, False),
        'C110': (validar_c110, False),
        'C170': (validar_c170, False),
        'C190': (validar_c190, False),
    }
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    for codigo_registro, (funcao_validar, apenas_primeira) in validadores.items():
        linhas_registro = [linha for linha in linhas if linha.startswith(f'|{codigo_registro}|')]
        
        if not linhas_registro:
            print(f"\nNenhuma linha do registro {codigo_registro} encontrada. Pulando...")
            continue
        
        print(f"\n{'=' * 60}")
        print(f"Processando Registro {codigo_registro} ({len(linhas_registro)} ocorrências)")
        print(f"{'=' * 60}")
        
        try:
            if apenas_primeira:
                if linhas_registro:
                    json_resultado = funcao_validar(linhas_registro[0])
                else:
                    json_resultado = None
            else:
                json_resultado = funcao_validar(linhas_registro)
            
            if json_resultado:
                nome_arquivo = base_dir / f"registro_{codigo_registro}_{timestamp}.xlsx"
                criar_arquivo_excel_registro_json(
                    json_resultado,
                    codigo_registro,
                    str(nome_arquivo),
                    dados_bloco_0,
                    participantes,
                    linhas
                )
            else:
                print(f"  Nenhum registro {codigo_registro} válido encontrado")
        except Exception as e:
            print(f"  Erro ao processar registro {codigo_registro}: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'=' * 60}")
    print("Processamento concluído!")
    print(f"{'=' * 60}")
    

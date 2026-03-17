import os
import re
import json
from datetime import datetime
from collections import OrderedDict
from pathlib import Path
from utils.cpf import formatar_cpf
from utils.cnpj import formatar_cnpj
from utils.data_hora import formatar_data
from utils.moeda import formatar_valor_monetario

try:
    from openpyxl import Workbook
except ImportError:
    import subprocess
    import sys
    print("AVISO: Biblioteca openpyxl não encontrada. Instalando...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook

from parser_bloco_0 import ler_bloco_0_do_sped, extrair_datas_do_registro_0000


def normalizar_partes_sped(partes):
    """
    Normaliza as partes de uma linha SPED após split por pipes.
    Substitui strings vazias (campos vazios entre pipes consecutivos ||) por "-".
    
    Args:
        partes: Lista de strings resultante de split('|')
        
    Returns:
        Lista normalizada com strings vazias substituídas por "-"
    """
    if not partes:
        return partes

    if partes and not partes[0]:
        partes = partes[1:]

    if partes and not partes[-1]:
        partes = partes[:-1]

    partes_normalizadas = [parte if parte.strip() else "-" for parte in partes]
    
    return partes_normalizadas


def ler_definicoes_registros(arquivo_markdown):
    """
    Lê o arquivo markdown e extrai as definições de todos os registros.
    
    Args:
        arquivo_markdown: Caminho para o arquivo sped_bloco_c.md
        
    Returns:
        Dicionário com código do registro como chave e lista de campos como valor
        Exemplo: {'C001': [{'numero': '01', 'campo': 'REG', 'tipo': 'C', ...}, ...]}
    """
    definicoes = OrderedDict()
    
    if not os.path.exists(arquivo_markdown):
        print(f"Erro: Arquivo '{arquivo_markdown}' não encontrado!")
        return definicoes
    
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(arquivo_markdown, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            print(f"Encoding do markdown detectado: {encoding}")
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo markdown!")
        return definicoes
    
    try:
        conteudo = arquivo.read()
    finally:
        arquivo.close()
    
    padrao_registro = r'(?:##\s+)?REGISTRO\s+(C\d+):'
    
    matches = list(re.finditer(padrao_registro, conteudo, re.IGNORECASE))
    
    for i, match in enumerate(matches):
        codigo_registro = match.group(1).upper()
        inicio = match.end()
        
        if i + 1 < len(matches):
            fim = matches[i + 1].start()
        else:
            fim = len(conteudo)
        
        secao = conteudo[inicio:fim]

        padrao_cabecalho = r'Nº\s+Campo\s+Descrição\s+Tipo\s+Tam\s+Dec'
        match_cabecalho = re.search(padrao_cabecalho, secao, re.IGNORECASE)
        
        if match_cabecalho:
            inicio_tabela = match_cabecalho.end()
            linhas = secao[inicio_tabela:].split('\n')
            
            campos = []
            campo_em_construcao = None
            
            for idx, linha in enumerate(linhas):
                linha_original = linha
                linha = linha.strip()

                if (linha.startswith('Observações') or linha.startswith('Observacoes') or 
                    linha.startswith('Nível') or linha.startswith('Nivel') or
                    (linha.startswith('Campo') and not campo_em_construcao) or 
                    linha.startswith('Validação') or
                    linha.startswith('REGISTRO') or linha.startswith('Guia Prático') or
                    linha.startswith('Página') or linha.startswith('---') or
                    (linha.startswith('###') or linha.startswith('##'))):
                    if campos and not campo_em_construcao:
                        break
                    if campo_em_construcao:
                        campos.append(campo_em_construcao)
                        campo_em_construcao = None
                    continue
                
                if not linha:
                    if campo_em_construcao:
                        campo_em_construcao['descricao'] += ' '
                    continue

                if linha.startswith('|') and linha.count('|') >= 3:
                    partes = [p.strip() for p in linha.split('|')]
                    
                    if partes and not partes[0]:
                        partes = partes[1:]
                    
                    if partes and not partes[-1]:
                        partes = partes[:-1]
                    
                    if len(partes) >= 6:
                        try:
                            numero_campo = partes[0].strip()
                            if not re.match(r'^\d{2}$', numero_campo):
                                continue
                            
                            nome_campo = partes[1].strip()
                            descricao = partes[2].strip()
                            tipo = partes[3].strip()
                            tamanho = partes[4].strip()
                            decimais = partes[5].strip()
                            obrigatorio = partes[6].strip() if len(partes) > 6 else ''
                            
                            if obrigatorio:
                                obrigatorio = obrigatorio.split()[0] if obrigatorio.split() else obrigatorio
                            
                            descricao = re.sub(r'<br\s*/?>', ' ', descricao, flags=re.IGNORECASE)
                            descricao = re.sub(r'<[^>]+>', '', descricao)
                            descricao = descricao.strip()
                            
                            campo = {
                                'numero': numero_campo,
                                'campo': nome_campo,
                                'descricao': descricao,
                                'tipo': tipo,
                                'tamanho': tamanho,
                                'decimais': decimais,
                                'obrigatorio': obrigatorio
                            }
                            campos.append(campo)
                            campo_em_construcao = None
                        except (IndexError, ValueError):
                            continue
                    continue

                match_novo_campo = re.match(r'^(\d{2})\s+(\w+)', linha)
                
                if match_novo_campo:
                    if campo_em_construcao:
                        desc = campo_em_construcao['descricao'].strip()
                        match_final = re.search(r'([CN])\s+(\d+|\d+\*|\-)\s+(\-|\d+)\s+([OC])', desc)
                        if match_final:
                            campo_em_construcao['tipo'] = match_final.group(1)
                            campo_em_construcao['tamanho'] = match_final.group(2)
                            campo_em_construcao['decimais'] = match_final.group(3)
                            campo_em_construcao['obrigatorio'] = match_final.group(4)
                            campo_em_construcao['descricao'] = desc[:match_final.start()].strip()
                        campos.append(campo_em_construcao)
                    
                    numero_campo = match_novo_campo.group(1)
                    nome_campo = match_novo_campo.group(2)
                    resto_linha = linha[len(match_novo_campo.group(0)):].strip()

                    match_completo = re.search(r'([CN])\s+(\d+|\d+\*|\-)\s+(\-|\d+)\s+([OC])', resto_linha)
                    
                    if match_completo:
                        tipo = match_completo.group(1)
                        tamanho = match_completo.group(2)
                        decimais = match_completo.group(3)
                        obrigatorio = match_completo.group(4)
                        descricao = resto_linha[:match_completo.start()].strip()
                        
                        descricao = re.sub(r'<br\s*/?>', ' ', descricao, flags=re.IGNORECASE)
                        descricao = re.sub(r'<[^>]+>', '', descricao)
                        
                        campo = {
                            'numero': numero_campo,
                            'campo': nome_campo,
                            'descricao': descricao,
                            'tipo': tipo,
                            'tamanho': tamanho,
                            'decimais': decimais,
                            'obrigatorio': obrigatorio
                        }
                        campos.append(campo)
                        campo_em_construcao = None
                    else:
                        descricao = resto_linha
                        campo_em_construcao = {
                            'numero': numero_campo,
                            'campo': nome_campo,
                            'descricao': descricao,
                            'tipo': '',
                            'tamanho': '',
                            'decimais': '',
                            'obrigatorio': ''
                        }
                else:
                    if campo_em_construcao:
                        match_final = re.search(r'([CN])\s+(\d+|\d+\*|\-)\s+(\-|\d+)\s+([OC])', linha)
                        if match_final:
                            campo_em_construcao['tipo'] = match_final.group(1)
                            campo_em_construcao['tamanho'] = match_final.group(2)
                            campo_em_construcao['decimais'] = match_final.group(3)
                            campo_em_construcao['obrigatorio'] = match_final.group(4)
                            
                            parte_desc = linha[:match_final.start()].strip()
                            if parte_desc:
                                campo_em_construcao['descricao'] += ' ' + parte_desc
                            campo_em_construcao['descricao'] = campo_em_construcao['descricao'].strip()
                            
                            campo_em_construcao['descricao'] = re.sub(r'<br\s*/?>', ' ', campo_em_construcao['descricao'], flags=re.IGNORECASE)
                            campo_em_construcao['descricao'] = re.sub(r'<[^>]+>', '', campo_em_construcao['descricao'])
                            campos.append(campo_em_construcao)
                            campo_em_construcao = None
                        else:
                            campo_em_construcao['descricao'] += ' ' + linha

            if campo_em_construcao:
                desc = campo_em_construcao['descricao'].strip()
                match_final = re.search(r'([CN])\s+(\d+|\d+\*|\-)\s+(\-|\d+)\s+([OC])', desc)
                if match_final:
                    campo_em_construcao['tipo'] = match_final.group(1)
                    campo_em_construcao['tamanho'] = match_final.group(2)
                    campo_em_construcao['decimais'] = match_final.group(3)
                    campo_em_construcao['obrigatorio'] = match_final.group(4)
                    campo_em_construcao['descricao'] = desc[:match_final.start()].strip()
                campos.append(campo_em_construcao)
            
            if campos:
                campos_limpos = []
                for campo in campos:
                    if campo.get('tipo') and campo.get('tamanho'):
                        campos_limpos.append(campo)
                    elif campo.get('campo'):
                        print(f"  AVISO: Campo {campo.get('campo')} do registro {codigo_registro} está incompleto (sem tipo/tamanho)")
                
                if campos_limpos:
                    numeros_campos = [int(c['numero']) for c in campos_limpos]
                    numeros_campos.sort()
                    if numeros_campos:
                        esperado = list(range(1, max(numeros_campos) + 1))
                        faltantes = [n for n in esperado if n not in numeros_campos]
                        if faltantes:
                            print(f"  AVISO: Registro {codigo_registro} tem campos faltantes na sequência: {faltantes}")
                    
                    definicoes[codigo_registro] = campos_limpos
                    print(f"Registro {codigo_registro} encontrado com {len(campos_limpos)} campos")
                else:
                    print(f"  AVISO: Registro {codigo_registro} não possui campos válidos definidos")
            else:
                print(f"  AVISO: Registro {codigo_registro} não possui campos definidos na tabela")
    
    print(f"\nTotal de registros encontrados: {len(definicoes)}")
    return definicoes


def parse_registro_generico(linha, codigo_registro, campos_definicao):
    """
    Parseia uma linha do SPED de forma genérica baseado nas definições.
    
    Args:
        linha: String com a linha do arquivo SPED
        codigo_registro: Código do registro (ex: 'C001')
        campos_definicao: Lista de dicionários com as definições dos campos
        
    Returns:
        Dicionário com os campos parseados ou None se inválido
    """
    linha = linha.strip()
    if not linha or codigo_registro.upper() not in linha:
        return None
    
    partes = linha.split('|')
    partes = normalizar_partes_sped(partes)
    
    try:
        idx = None
        for i, p in enumerate(partes):
            if p.strip().upper() == codigo_registro.upper():
                idx = i
                break
        
        if idx is None:
            return None
        
        campos = {}
        for campo_def in campos_definicao:
            numero_campo = int(campo_def['numero'])
            nome_campo = campo_def['campo']
            
            posicao = idx + (numero_campo - 1)
            
            if posicao < len(partes):
                valor = partes[posicao].strip()
                campos[nome_campo] = valor if valor else '-'
            else:
                campos[nome_campo] = '-'
        
        return campos
        
    except (IndexError, ValueError, KeyError) as e:
        print(f"Erro ao parsear linha do registro {codigo_registro}: {e}")
        return None


def ler_dados_participantes():
    """
    Lê os dados do registro 0150 do bloco 0 para obter informações dos participantes.
    
    Returns:
        Dicionário com COD_PART como chave e dicionário com dados do participante como valor
        Exemplo: {'PART001': {'cod_part': 'PART001', 'cnpj': '12.345.678/0001-90', 'cpf': '', 'nome': 'Empresa XYZ'}}
    """
    participantes = {}
    
    base_dir = Path(__file__).parent
    
    caminho_0150 = base_dir / "blocos_separados" / "bloco_0" / "0150.txt"
    if not caminho_0150.exists():
        return participantes
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_0150, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        return participantes
    
    try:
        for linha in arquivo:
            linha = linha.strip()
            if linha and linha.startswith('|'):
                partes = linha.split('|')
                partes = normalizar_partes_sped(partes)
                
                if len(partes) > 0 and partes[0].strip() == '0150':
                    if len(partes) >= 6:
                        cod_part = partes[1].strip() if len(partes) > 1 else ''
                        nome = partes[2].strip() if len(partes) > 2 else ''
                        cnpj = partes[4].strip() if len(partes) > 4 else ''
                        cpf = partes[5].strip() if len(partes) > 5 else ''

                        if cod_part == '-':
                            cod_part = ''
                        if nome == '-':
                            nome = ''
                        if cnpj == '-':
                            cnpj = ''
                        if cpf == '-':
                            cpf = ''
                        
                        if cod_part:
                            cnpj_formatado = formatar_cnpj(cnpj) if cnpj and len(cnpj) == 14 else cnpj
                            cpf_formatado = formatar_cpf(cpf) if cpf and len(cpf) == 11 else cpf         
                            participantes[cod_part] = {
                                'cod_part': cod_part,
                                'cnpj': cnpj_formatado,
                                'cpf': cpf_formatado,
                                'nome': nome
                            }
    finally:
        arquivo.close()
    
    return participantes


def obter_descricao_campo(campo, valor):
    """
    Retorna descrição formatada para campos específicos baseado no valor.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        
    Returns:
        String com descrição formatada ou valor original
    """
    if not valor:
        return valor
    
    descricoes = {
        'IND_MOV': {
            '0': '0 - Bloco com dados informados',
            '1': '1 - Bloco sem dados informados'
        },
        'IND_OPER': {
            '0': '0 - Entrada',
            '1': '1 - Saída'
        },
        'IND_EMIT': {
            '0': '0 - Emissão própria',
            '1': '1 - Terceiros'
        },
        'IND_PGTO': {
            '0': '0 - À vista',
            '1': '1 - A prazo',
            '2': '2 - Sem indicação',
            '9': '9 - Sem pagamento'
        },
        'IND_FRT': {
            '0': '0 - Por conta do remetente (CIF)',
            '1': '1 - Por conta do destinatário (FOB)',
            '2': '2 - Por conta de terceiros',
            '9': '9 - Sem frete'
        }
    }
    
    if campo in descricoes and valor in descricoes[campo]:
        return descricoes[campo][valor]
    
    return valor


def aplicar_formatacao(campo, valor, tipo_campo):
    """
    Aplica formatação específica baseada no tipo e nome do campo.
    
    Args:
        campo: Nome do campo
        valor: Valor do campo
        tipo_campo: Tipo do campo (C, N, etc)
        
    Returns:
        Valor formatado
    """
    if not valor:
        return valor
    
    if campo == 'CNPJ' and len(valor) == 14:
        return formatar_cnpj(valor)
    
    if campo.startswith('DT_') or campo.startswith('DTE_'):
        if len(valor) == 8 and valor.isdigit():
            return formatar_data(valor)
    
    if campo.startswith('VL_') and tipo_campo == 'N':
        return formatar_valor_monetario(valor)
    
    return obter_descricao_campo(campo, valor)


def extrair_dados_do_registro_0000(linhas_bloco_0):
    """
    Extrai os dados do registro 0000 das linhas do bloco 0.
    
    Args:
        linhas_bloco_0: Lista de linhas do bloco 0
        
    Returns:
        Dicionário com cnpj, ie e periodo, ou valores vazios se não encontrados
    """
    dados = {
        'cnpj': '',
        'ie': '',
        'periodo': ''
    }

    for linha in linhas_bloco_0:
        if '|0000|' in linha or linha.startswith('|0000|'):
            partes = linha.split('|')

            if len(partes) >= 6:
                if partes[1].strip() == '0000':
                    if len(partes) >= 11:
                        cnpj = partes[7].strip() if len(partes) > 7 else ''
                        if cnpj and cnpj != '-':
                            dados['cnpj'] = formatar_cnpj(cnpj) if len(cnpj) == 14 else cnpj

                        dt_ini = partes[4].strip() if len(partes) > 4 else ''
                        dt_fin = partes[5].strip() if len(partes) > 5 else ''
                        if dt_ini and dt_ini != '-' and dt_fin and dt_fin != '-' and len(dt_ini) == 8 and len(dt_fin) == 8:
                            periodo_ini = formatar_data(dt_ini)
                            periodo_fin = formatar_data(dt_fin)
                            dados['periodo'] = f"{periodo_ini} a {periodo_fin}"

                        ie = partes[10].strip() if len(partes) > 10 else ''
                        if ie and ie != '-':
                            dados['ie'] = ie
                    
                    break
    
    return dados


def ler_dados_bloco_0(linhas_bloco_0=None):
    """
    Lê os dados do bloco 0 necessários para as colunas iniciais.
    
    Args:
        linhas_bloco_0: Lista opcional de linhas do bloco 0. Se não fornecida, tenta ler de arquivo separado.
    
    Returns:
        Dicionário com cnpj, ie e periodo, ou valores vazios se não encontrados
    """
    # Se linhas do bloco 0 foram fornecidas, usa elas
    if linhas_bloco_0:
        return extrair_dados_do_registro_0000(linhas_bloco_0)
    
    # Caso contrário, tenta ler de arquivo separado (comportamento antigo)
    dados = {
        'cnpj': '',
        'ie': '',
        'periodo': ''
    }
    
    base_dir = Path(__file__).parent
    
    caminho_0000 = base_dir / "blocos_separados" / "bloco_0" / "0000.txt"
    if caminho_0000.exists():
        encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
        arquivo = None
        
        for encoding in encodings:
            try:
                arquivo = open(caminho_0000, 'r', encoding=encoding)
                arquivo.readline()
                arquivo.seek(0)
                break
            except (UnicodeDecodeError, UnicodeError):
                if arquivo:
                    arquivo.close()
                continue
        
        if arquivo:
            try:
                for linha in arquivo:
                    linha = linha.strip()
                    if linha and linha.startswith('|'):
                        partes = linha.split('|')
                        partes = normalizar_partes_sped(partes)
                        if len(partes) > 0 and partes[0].strip() == '0000':
                            if len(partes) >= 11:
                                cnpj = partes[7].strip() if len(partes) > 7 else ''
                                dt_ini = partes[4].strip() if len(partes) > 4 else ''
                                dt_fin = partes[5].strip() if len(partes) > 5 else ''
                                
                                if cnpj and cnpj != '-':
                                    dados['cnpj'] = formatar_cnpj(cnpj) if len(cnpj) == 14 else cnpj
                                
                                if dt_ini and dt_ini != '-' and dt_fin and dt_fin != '-':
                                    if len(dt_ini) == 8 and len(dt_fin) == 8:
                                        periodo_ini = formatar_data(dt_ini)
                                        periodo_fin = formatar_data(dt_fin)
                                        dados['periodo'] = f"{periodo_ini} a {periodo_fin}"
                            break
            finally:
                arquivo.close()
    
    if not dados['ie'] and caminho_0000.exists():
        encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
        arquivo = None
        
        for encoding in encodings:
            try:
                arquivo = open(caminho_0000, 'r', encoding=encoding)
                arquivo.readline()
                arquivo.seek(0)
                break
            except (UnicodeDecodeError, UnicodeError):
                if arquivo:
                    arquivo.close()
                continue
        
        if arquivo:
            try:
                for linha in arquivo:
                    linha = linha.strip()
                    if linha and linha.startswith('|'):
                        partes = linha.split('|')
                        partes = normalizar_partes_sped(partes)

                        if len(partes) > 0 and partes[0].strip() == '0000':
                            if len(partes) >= 11:
                                ie = partes[10].strip() if len(partes) > 10 else ''

                                if ie and ie != '-':
                                    dados['ie'] = ie
                            break
            finally:
                arquivo.close()
    
    return dados


def criar_arquivo_excel_registro(registros, codigo_registro, campos_definicao, nome_arquivo=None, dados_bloco_0=None, participantes=None):
    """
    Cria um arquivo Excel (.xlsx) com os registros parseados.
    
    Args:
        registros: Lista de dicionários com os registros
        codigo_registro: Código do registro (ex: 'C001')
        campos_definicao: Lista de dicionários com as definições dos campos
        nome_arquivo: Nome do arquivo Excel a ser criado (opcional)
        dados_bloco_0: Dicionário com cnpj, ie e periodo do bloco 0 (opcional)
        participantes: Dicionário com dados dos participantes (opcional)
    """
    if not registros:
        print(f"Nenhum registro {codigo_registro} encontrado!")
        return None
    
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"registro_{codigo_registro}_{timestamp}.xlsx"
    
    if dados_bloco_0 is None:
        dados_bloco_0 = ler_dados_bloco_0()
    
    if participantes is None:
        participantes = ler_dados_participantes()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Registro {codigo_registro}"
    
    cabecalhos = ['CNPJ', 'Inscrição Estadual', 'Período']
    cabecalhos.extend(['Código Participante', 'CNPJ Participante', 'CPF Participante', 'Nome Participante'])
    cabecalhos.extend([campo['campo'] for campo in campos_definicao])
    
    for col_idx, cabecalho in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cabecalho)
    
    linha_atual = 2
    
    for registro in registros:
        valores = []
        valores.append(dados_bloco_0.get('cnpj', ''))
        valores.append(dados_bloco_0.get('ie', ''))
        valores.append(dados_bloco_0.get('periodo', ''))
        
        cod_part = registro.get('COD_PART', '')
        if cod_part and cod_part in participantes:
            participante = participantes[cod_part]
            valores.append(participante.get('cod_part', ''))
            valores.append(participante.get('cnpj', ''))
            valores.append(participante.get('cpf', ''))
            valores.append(participante.get('nome', ''))
        else:
            valores.extend(['', '', '', ''])
        
        for campo_def in campos_definicao:
            nome_campo = campo_def['campo']
            tipo_campo = campo_def['tipo']
            valor = registro.get(nome_campo, '')
            
            valor_formatado = aplicar_formatacao(nome_campo, valor, tipo_campo)
            valores.append(valor_formatado if valor_formatado else '')
        
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=linha_atual, column=col_idx, value=valor)
        
        linha_atual += 1
    
    try:
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        print(f"Total de registros processados: {len(registros)}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")
        return None


def ler_bloco_c_do_sped(caminho_arquivo_sped):
    """
    Lê o arquivo SPED e extrai todas as linhas do bloco C.
    Retorna uma lista de linhas do bloco C.
    """
    linhas_bloco_c = []
    
    if not os.path.exists(caminho_arquivo_sped):
        print(f"Erro: Arquivo SPED '{caminho_arquivo_sped}' não encontrado!")
        return linhas_bloco_c
    
    encodings = ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']
    arquivo = None
    
    for encoding in encodings:
        try:
            arquivo = open(caminho_arquivo_sped, 'r', encoding=encoding)
            arquivo.readline()
            arquivo.seek(0)
            break
        except (UnicodeDecodeError, UnicodeError):
            if arquivo:
                arquivo.close()
            continue
    
    if arquivo is None:
        print(f"Erro: Não foi possível determinar o encoding do arquivo SPED!")
        return linhas_bloco_c
    
    try:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            
            if linha.startswith('|') and '|' in linha[1:]:
                partes = linha.split('|')
                partes = normalizar_partes_sped(partes)

                if len(partes) > 0 and partes[0] and partes[0].upper().startswith('C'):
                    linhas_bloco_c.append(linha)
    finally:
        if arquivo:
            arquivo.close()
    
    return linhas_bloco_c


def extrair_mes_ano_da_data(data_ddmmyyyy):
    """
    Extrai mês e ano de uma data no formato DDMMYYYY.
    Retorna (ano, mes) como inteiros.
    """
    if not data_ddmmyyyy or len(data_ddmmyyyy) != 8:
        return None, None
    
    try:
        ano = int(data_ddmmyyyy[4:8])
        mes = int(data_ddmmyyyy[2:4])
        return ano, mes
    except:
        return None, None


def extrair_mes_ano_da_data_criacao(data_criacao):
    """
    Extrai mês e ano de uma data no formato YYYY-MM-DD.
    Retorna (ano, mes) como inteiros.
    """
    if not data_criacao:
        return None, None
    
    try:
        partes = data_criacao.split('-')
        if len(partes) >= 2:
            ano = int(partes[0])
            mes = int(partes[1])
            return ano, mes
    except:
        pass
    
    return None, None


def buscar_versao_por_mes_ano(ano_arquivo, mes_arquivo, caminho_versoes_json):
    """
    Busca a versão apropriada no arquivo versoes.json baseado no mês/ano.
    
    Lógica:
    1. Busca versão com mês/ano igual ao do arquivo
    2. Se não encontrar, busca o mês mais avançado (anterior) ao período do arquivo
       (exemplo: arquivo 02/2026, só há 01/2026 e 03/2026, usa 01/2026)
    
    Retorna a versão encontrada ou None.
    """
    if not os.path.exists(caminho_versoes_json):
        print(f"Erro: Arquivo '{caminho_versoes_json}' não encontrado!")
        return None
    
    with open(caminho_versoes_json, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    versoes = dados.get("versoes", [])
    
    if not versoes:
        print("Erro: Nenhuma versão encontrada no arquivo versoes.json!")
        return None
    
    versao_exata = None
    versoes_anteriores = []
    
    for versao in versoes:
        data_criacao = versao.get("data_criacao", "")
        ano_versao, mes_versao = extrair_mes_ano_da_data_criacao(data_criacao)
        
        if ano_versao is None or mes_versao is None:
            continue
        
        if ano_versao == ano_arquivo and mes_versao == mes_arquivo:
            versao_exata = versao
            break

        if (ano_versao < ano_arquivo) or (ano_versao == ano_arquivo and mes_versao < mes_arquivo):
            versoes_anteriores.append((versao, ano_versao, mes_versao))
    
    if versao_exata:
        print(f"Versão encontrada (mês/ano exato): {versao_exata['versao']} (data_criacao: {versao_exata['data_criacao']})")
        return versao_exata

    if versoes_anteriores:
        versoes_anteriores.sort(key=lambda x: (x[1], x[2]), reverse=True)
        versao_encontrada = versoes_anteriores[0][0]
        print(f"Versão encontrada (mês anterior mais próximo): {versao_encontrada['versao']} (data_criacao: {versao_encontrada['data_criacao']})")
        print(f"  Período do arquivo: {mes_arquivo:02d}/{ano_arquivo}")
        return versao_encontrada

    versoes_ordenadas = sorted(versoes, key=lambda x: x.get("data_criacao", ""), reverse=True)
    if versoes_ordenadas:
        versao_encontrada = versoes_ordenadas[0]
        print(f"Aviso: Nenhuma versão anterior encontrada. Usando a mais recente: {versao_encontrada['versao']} (data_criacao: {versao_encontrada['data_criacao']})")
        return versao_encontrada
    
    return None


def processar_bloco_c(sped_arquivo=None):
    """
    Função principal que processa o bloco C do arquivo SPED.
    
    Fluxo:
    1. Lê o arquivo SPED e extrai linhas do bloco C
    2. Identifica DT_INI e DT_FIN no registro 0000 (do bloco 0)
    3. Busca a versão apropriada no versoes.json (por mês/ano)
    4. Carrega a documentação markdown da versão encontrada
    5. Processa todos os registros do bloco C
    6. Gera arquivos Excel para cada tipo de registro
    """
    base_dir = Path(__file__).parent
    
    if sped_arquivo is None:
        sped_arquivo = "08589276000169-9084073269-20201101-20201130-0-911F0BE8FE2F193E99CEFD2FA579B4E7F40C8ED9-SPED-EFD.txt"
    
    caminho_sped = base_dir / sped_arquivo if not os.path.isabs(sped_arquivo) else Path(sped_arquivo)
    caminho_versoes_json = base_dir / "versoes.json"
    
    print("=" * 60)
    print("Parser do Bloco C - SPED EFD Fiscal")
    print("=" * 60)
    print(f"\nLendo arquivo SPED: {caminho_sped}")

    linhas_bloco_0 = ler_bloco_0_do_sped(str(caminho_sped))
    dt_ini_str, dt_fin_str = extrair_datas_do_registro_0000(linhas_bloco_0)
    
    if not dt_ini_str or not dt_fin_str:
        print("Erro: Não foi possível extrair as datas DT_INI e DT_FIN do registro 0000!")
        return
    
    print(f"\nDatas extraídas do registro 0000:")
    print(f"  DT_INI: {dt_ini_str} (DDMMYYYY)")
    print(f"  DT_FIN: {dt_fin_str} (DDMMYYYY)")
    
    ano_arquivo, mes_arquivo = extrair_mes_ano_da_data(dt_fin_str)
    
    if not ano_arquivo or not mes_arquivo:
        print("Erro: Não foi possível extrair mês/ano das datas!")
        return
    
    print(f"  Período do arquivo: {mes_arquivo:02d}/{ano_arquivo}")
    
    print(f"\nBuscando versão correspondente em {caminho_versoes_json}...")
    versao_info = buscar_versao_por_mes_ano(ano_arquivo, mes_arquivo, str(caminho_versoes_json))
    
    if not versao_info:
        print("Erro: Não foi possível encontrar uma versão válida!")
        return
    
    versao = versao_info["versao"]
    
    caminho_markdown = base_dir / "documentacao_blocos" / versao / "bloco_C.md"
    
    if not caminho_markdown.exists():
        print(f"Erro: Arquivo de documentação não encontrado: {caminho_markdown}")
        return
    
    print(f"\nUsando documentação: {caminho_markdown}")
    print(f"Lendo definições dos registros...")
    
    definicoes = ler_definicoes_registros(str(caminho_markdown))
    
    if not definicoes:
        print("\nErro: Nenhuma definição de registro encontrada no arquivo markdown!")
        return
    
    print(f"\n{len(definicoes)} registros encontrados no markdown")
    
    print(f"\nLendo bloco C do arquivo SPED...")
    linhas_bloco_c = ler_bloco_c_do_sped(str(caminho_sped))
    
    if not linhas_bloco_c:
        print("Aviso: Nenhuma linha do bloco C encontrada no arquivo SPED!")
        return
    
    print(f"Encontradas {len(linhas_bloco_c)} linhas do bloco C")
    
    print("\nLendo dados do bloco 0...")

    dados_bloco_0 = ler_dados_bloco_0(linhas_bloco_0)
    if dados_bloco_0['cnpj']:
        print(f"CNPJ encontrado: {dados_bloco_0['cnpj']}")
    if dados_bloco_0['ie']:
        print(f"Inscrição Estadual encontrada: {dados_bloco_0['ie']}")
    if dados_bloco_0['periodo']:
        print(f"Período encontrado: {dados_bloco_0['periodo']}")
    
    print("\nLendo dados dos participantes (registro 0150)...")
    participantes = ler_dados_participantes()
    print(f"Total de participantes encontrados: {len(participantes)}")
    
    print("\nProcessando registros do bloco C...")
    print("=" * 60)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    registros_processados = {}
    
    for linha in linhas_bloco_c:
        partes = linha.split('|')
        partes = normalizar_partes_sped(partes)
        
        if len(partes) > 0:
            codigo_registro = partes[0].strip().upper()
            
            if codigo_registro and codigo_registro != '-':
                if codigo_registro not in registros_processados:
                    registros_processados[codigo_registro] = []
                
                registros_processados[codigo_registro].append(linha)
    
    for codigo_registro, linhas_registro in registros_processados.items():
        if codigo_registro not in definicoes:
            print(f"\nAviso: Registro {codigo_registro} não encontrado nas definições do markdown. Pulando...")
            continue
        
        print(f"\nProcessando Registro {codigo_registro} ({len(linhas_registro)} ocorrências)")
        
        campos_definicao = definicoes[codigo_registro]
        
        registros_parseados = []
        for linha in linhas_registro:
            registro = parse_registro_generico(linha, codigo_registro, campos_definicao)
            if registro:
                registros_parseados.append(registro)
        
        if registros_parseados:
            nome_arquivo_excel = base_dir / f"registro_{codigo_registro}_{timestamp}.xlsx"
            criar_arquivo_excel_registro(registros_parseados, codigo_registro, campos_definicao, str(nome_arquivo_excel), dados_bloco_0, participantes)
        else:
            print(f"  Nenhum registro {codigo_registro} válido encontrado")
    
    print("\n" + "=" * 60)
    print("Processamento concluído!")
    print("=" * 60)


if __name__ == "__main__":
    sped_arquivo = "38045115000841-0390189812-20250201-20250228-0-769A2231EC07CDAA66A10B417EC8F709B8A1894D-SPED-EFD.txt"
    processar_bloco_c(sped_arquivo)
